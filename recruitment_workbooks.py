"""Excel and CSV outputs for the recruitment ground-truth pipeline."""

from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Any, Mapping, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from exports import dataframe_to_csv_bytes
from recruitment_config import (
    CATEGORIES,
    CATEGORY_LABELS,
    MONTH_LABELS,
    ProjectConfig,
)


TITLE_GREY = "7B7F86"
BLUE = "6697C5"
CURRENT_ORANGE = "FFB12B"
LIGHT_GREY = "EEF0F3"
SPACER_GREY = "C7C7C7"
GRID_GREY = "AAB3BE"
WHITE = "FFFFFF"
BLACK = "000000"
GREEN = "2E7D32"
RED = "C0392B"
AMBER = "F4B183"
PALE_GREEN = "E2F0D9"
PALE_RED = "FCE4D6"

THIN_BORDER = Border(
    left=Side(style="thin", color=GRID_GREY),
    right=Side(style="thin", color=GRID_GREY),
    top=Side(style="thin", color=GRID_GREY),
    bottom=Side(style="thin", color=GRID_GREY),
)


def _excel_value(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _current_index(milestones: Sequence[dt.date], report_date: dt.date) -> int:
    return next(
        (
            index
            for index, milestone in enumerate(milestones)
            if milestone >= report_date
        ),
        len(milestones) - 1,
    )


def _project_title(project_key: str, config: ProjectConfig) -> str:
    if config.grant:
        return (
            f"Recruitment Milestones for {config.grant} - {config.study_title}"
        )
    return f"Recruitment Milestones for {config.study_title}"


def _year_spans(milestones: Sequence[dt.date]) -> list[tuple[int, int, int]]:
    spans: list[tuple[int, int, int]] = []
    start = 0
    while start < len(milestones):
        year = milestones[start].year
        stop = start
        while stop + 1 < len(milestones) and milestones[stop + 1].year == year:
            stop += 1
        spans.append((year, start, stop))
        start = stop + 1
    return spans


def _set_cell_style(
    cell: Any,
    *,
    fill: str | None = None,
    color: str = BLACK,
    bold: bool = False,
    horizontal: str = "center",
    vertical: str = "center",
    wrap: bool = False,
) -> None:
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(
        name="Arial",
        size=10,
        bold=bold,
        color=color,
    )
    cell.alignment = Alignment(
        horizontal=horizontal,
        vertical=vertical,
        wrap_text=wrap,
    )
    cell.border = THIN_BORDER


def _write_milestone_summary(
    *,
    worksheet: Any,
    project_key: str,
    config: ProjectConfig,
    milestones: Sequence[dt.date],
    actuals: Mapping[str, Sequence[int | None]],
    report_date: dt.date,
    project_info: Mapping[str, Any],
    inventory: Mapping[str, Any],
    access_evidence: Mapping[str, Any],
) -> None:
    """Write the image-matched milestone table and source evidence."""

    last_column = len(milestones) + 1
    current_index = _current_index(milestones, report_date)
    current_column = current_index + 2

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "B5"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.print_title_rows = "1:4"

    worksheet.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=last_column
    )
    title = worksheet.cell(1, 1, _project_title(project_key, config))
    _set_cell_style(title, fill=TITLE_GREY, color=WHITE, bold=True)
    title.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    worksheet.row_dimensions[1].height = 20

    worksheet.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=last_column
    )
    mapping_note = (
        "Source status: no protocol-confirmed enrollment status/date mapping is "
        "present; diagnostic dates are not used. Missing values are N/A."
    )
    note = worksheet.cell(2, 1, mapping_note)
    _set_cell_style(
        note,
        fill=LIGHT_GREY,
        color=BLACK,
        bold=True,
        horizontal="left",
        wrap=True,
    )
    worksheet.row_dimensions[2].height = 34

    worksheet.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    milestone_label = worksheet.cell(3, 1, "Tri Yearly Milestones")
    _set_cell_style(
        milestone_label,
        fill=BLUE,
        color=WHITE,
        bold=True,
        horizontal="center",
    )

    for year, first_index, last_index in _year_spans(milestones):
        start_column = first_index + 2
        end_column = last_index + 2
        worksheet.merge_cells(
            start_row=3,
            start_column=start_column,
            end_row=3,
            end_column=end_column,
        )
        year_cell = worksheet.cell(3, start_column, year)
        _set_cell_style(year_cell, fill=BLUE, color=WHITE, bold=True)
        for column in range(start_column + 1, end_column + 1):
            _set_cell_style(
                worksheet.cell(3, column), fill=BLUE, color=WHITE, bold=True
            )

    for index, milestone in enumerate(milestones):
        column = index + 2
        month_cell = worksheet.cell(4, column, MONTH_LABELS[milestone.month])
        _set_cell_style(
            month_cell,
            fill=CURRENT_ORANGE if column == current_column else BLUE,
            color=BLACK if column == current_column else WHITE,
            bold=True,
        )
        month_cell.font = Font(
            name="Arial",
            size=10,
            bold=True,
            underline="single",
            color=BLACK if column == current_column else WHITE,
        )

    for column in range(1, last_column + 1):
        spacer = worksheet.cell(5, column)
        _set_cell_style(spacer, fill=SPACER_GREY)
    worksheet.row_dimensions[5].height = 10

    metric_rows: dict[tuple[str, str], int] = {}
    row = 6
    for category_index, category in enumerate(CATEGORIES):
        if category_index:
            for column in range(1, last_column + 1):
                spacer = worksheet.cell(row, column)
                _set_cell_style(spacer, fill=SPACER_GREY)
            worksheet.row_dimensions[row].height = 10
            row += 1

        metric_specs = (
            ("Previous Target", "previous"),
            ("Current Target", "current"),
            ("Actual", "actual"),
            ("Actual/Target Ratio", "ratio"),
            ("Status", "status"),
        )
        for metric_label, metric_key in metric_specs:
            metric_rows[(category, metric_key)] = row
            label = worksheet.cell(
                row,
                1,
                f"{metric_label}: {CATEGORY_LABELS[category]}",
            )
            _set_cell_style(
                label,
                fill=BLUE,
                color=WHITE,
                bold=True,
                horizontal="left",
            )
            for index, _milestone in enumerate(milestones):
                column = index + 2
                cell = worksheet.cell(row, column)
                base_fill = CURRENT_ORANGE if column == current_column else WHITE
                if metric_key == "previous" and column != current_column:
                    base_fill = LIGHT_GREY

                if metric_key == "previous":
                    values = config.previous_targets.get(category, [])
                    cell.value = values[index] if index < len(values) else None
                elif metric_key == "current":
                    values = config.current_targets.get(category, [])
                    cell.value = values[index] if index < len(values) else None
                elif metric_key == "actual":
                    values = actuals.get(category, [])
                    cell.value = values[index] if index < len(values) else None
                elif metric_key == "ratio":
                    actual_row = metric_rows[(category, "actual")]
                    target_row = metric_rows[(category, "current")]
                    actual_ref = f"{get_column_letter(column)}{actual_row}"
                    target_ref = f"{get_column_letter(column)}{target_row}"
                    cell.value = (
                        f'=IF(OR({actual_ref}="",{target_ref}="",'
                        f'{target_ref}=0),"N/A",{actual_ref}/{target_ref})'
                    )
                    cell.number_format = "0%"
                elif metric_key == "status":
                    actual_row = metric_rows[(category, "actual")]
                    target_row = metric_rows[(category, "current")]
                    actual_ref = f"{get_column_letter(column)}{actual_row}"
                    target_ref = f"{get_column_letter(column)}{target_row}"
                    cell.value = (
                        f'=IF(OR({target_ref}="",{target_ref}=0),"N/A",'
                        f'IF({actual_ref}="","Not reported",'
                        f'IF({actual_ref}>={target_ref},"On target","Behind")))'
                    )
                _set_cell_style(
                    cell,
                    fill=base_fill,
                    bold=metric_key in {"previous", "current", "ratio", "status"},
                )
            row += 1

    table_last_row = row - 1
    for worksheet_row in worksheet.iter_rows(
        min_row=3, max_row=table_last_row, min_col=current_column, max_col=current_column
    ):
        for cell in worksheet_row:
            if cell.row not in (3, 4) and cell.fill.fgColor.rgb != "00C7C7C7":
                cell.fill = PatternFill("solid", fgColor=CURRENT_ORANGE)
                cell.font = Font(
                    name="Arial",
                    size=10,
                    bold=True,
                    color=BLACK,
                )

    status_ranges = [
        f"B{metric_rows[(category, 'status')]}:"
        f"{get_column_letter(last_column)}{metric_rows[(category, 'status')]}"
        for category in CATEGORIES
    ]
    for cell_range in status_ranges:
        worksheet.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'B{cell_range.split("B")[1].split(":")[0]}="On target"'],
                font=Font(color=GREEN, bold=True),
            ),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'B{cell_range.split("B")[1].split(":")[0]}="Behind"'],
                font=Font(color=RED, bold=True),
            ),
        )

    evidence_start = table_last_row + 3
    worksheet.merge_cells(
        start_row=evidence_start,
        start_column=1,
        end_row=evidence_start,
        end_column=last_column,
    )
    evidence_title = worksheet.cell(
        evidence_start, 1, "Source, access, and mapping evidence"
    )
    _set_cell_style(
        evidence_title, fill=TITLE_GREY, color=WHITE, bold=True, horizontal="left"
    )
    evidence_rows = [
        ("Project key", project_key),
        ("Observed REDCap project ID", project_info.get("project_id", "")),
        ("Observed REDCap project title", project_info.get("project_title", "")),
        ("Report cutoff", report_date.isoformat()),
        (
            "Enrollment-status mapping",
            config.enrollment_status_field
            or "N/A - no protocol-confirmed field/rule in repository evidence",
        ),
        (
            "Enrollment-date mapping",
            config.date_anchor
            or "N/A - no protocol-confirmed enrollment/consent date",
        ),
        ("Target provenance", config.target_provenance),
        ("Actual provenance", config.actual_provenance),
        (
            "Required API reads",
            ", ".join(
                key
                for key, value in access_evidence.items()
                if bool(value)
            )
            or "N/A",
        ),
        (
            "Participant records visible to token",
            inventory.get("participant_records", ""),
        ),
        (
            "Unresolved inclusion",
            inventory.get("unresolved_inclusion", ""),
        ),
        (
            "Definitely excluded by source flags",
            inventory.get("definitely_excluded", ""),
        ),
    ]
    for offset, (label_text, value) in enumerate(evidence_rows, start=1):
        evidence_row = evidence_start + offset
        worksheet.cell(evidence_row, 1, label_text)
        worksheet.merge_cells(
            start_row=evidence_row,
            start_column=2,
            end_row=evidence_row,
            end_column=last_column,
        )
        worksheet.cell(evidence_row, 2, _excel_value(value))
        _set_cell_style(
            worksheet.cell(evidence_row, 1),
            fill=BLUE,
            color=WHITE,
            bold=True,
            horizontal="left",
            wrap=True,
        )
        _set_cell_style(
            worksheet.cell(evidence_row, 2),
            fill=WHITE,
            horizontal="left",
            wrap=True,
        )
        for column in range(3, last_column + 1):
            _set_cell_style(
                worksheet.cell(evidence_row, column),
                fill=WHITE,
                horizontal="left",
                wrap=True,
            )
        worksheet.row_dimensions[evidence_row].height = 30

    worksheet.column_dimensions["A"].width = 54
    for column in range(2, last_column + 1):
        worksheet.column_dimensions[get_column_letter(column)].width = 13
    worksheet.auto_filter.ref = (
        f"A{evidence_start + 1}:{get_column_letter(last_column)}"
        f"{evidence_start + len(evidence_rows)}"
    )


def _write_dataframe_sheet(
    worksheet: Any,
    frame: pd.DataFrame,
    *,
    table_name: str,
    freeze: str = "A2",
) -> None:
    worksheet.sheet_view.showGridLines = False
    columns = [str(column) for column in frame.columns]
    for column_index, column in enumerate(columns, start=1):
        cell = worksheet.cell(1, column_index, column)
        _set_cell_style(cell, fill=BLUE, color=WHITE, bold=True, wrap=True)
    for row_index, row_values in enumerate(
        frame.itertuples(index=False, name=None), start=2
    ):
        for column_index, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row_index, column_index, _excel_value(value))
            _set_cell_style(
                cell,
                fill=WHITE,
                horizontal="left",
                vertical="top",
                wrap=True,
            )

    if columns:
        last_column = get_column_letter(len(columns))
        last_row = max(1, len(frame) + 1)
        worksheet.auto_filter.ref = f"A1:{last_column}{last_row}"
        if len(frame):
            table = Table(
                displayName=table_name,
                ref=f"A1:{last_column}{last_row}",
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)
        worksheet.freeze_panes = freeze

    for column_index, column in enumerate(columns, start=1):
        values = [
            str(column),
            *(str(value) for value in frame[column].head(200).tolist()),
        ]
        width = min(46, max(12, max(len(value) for value in values) + 2))
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def _style_audit_decisions(worksheet: Any, frame: pd.DataFrame) -> None:
    if frame.empty or "included_in_recruitment_count" not in frame.columns:
        return
    column_index = list(frame.columns).index("included_in_recruitment_count") + 1
    column_letter = get_column_letter(column_index)
    data_range = f"{column_letter}2:{column_letter}{len(frame) + 1}"
    worksheet.conditional_formatting.add(
        data_range,
        CellIsRule(
            operator="equal",
            formula=["TRUE"],
            fill=PatternFill("solid", fgColor=PALE_GREEN),
        ),
    )
    worksheet.conditional_formatting.add(
        data_range,
        CellIsRule(
            operator="equal",
            formula=["FALSE"],
            fill=PatternFill("solid", fgColor=PALE_RED),
        ),
    )
    worksheet.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'{column_letter}2=""'],
            fill=PatternFill("solid", fgColor=AMBER),
        ),
    )


def write_project_workbook(
    *,
    path: Path,
    project_key: str,
    config: ProjectConfig,
    milestones: Sequence[dt.date],
    actuals: Mapping[str, Sequence[int | None]],
    report_date: dt.date,
    project_info: Mapping[str, Any],
    inventory: Mapping[str, Any],
    access_evidence: Mapping[str, Any],
    participant_audit: pd.DataFrame | None = None,
    combined_milestone_summary: pd.DataFrame | None = None,
    data_quality_issues: pd.DataFrame | None = None,
) -> Path:
    """Write one project workbook; optional audit sheets make it restricted."""

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.properties.creator = "ESD REDCap recruitment pipeline"
    workbook.properties.title = f"{project_key} Recruitment Ground Truth"
    workbook.properties.subject = (
        "Restricted participant audit" if participant_audit is not None else "Aggregate"
    )
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    summary = workbook.active
    summary.title = f"{project_key}_Milestone_Summary"
    _write_milestone_summary(
        worksheet=summary,
        project_key=project_key,
        config=config,
        milestones=milestones,
        actuals=actuals,
        report_date=report_date,
        project_info=project_info,
        inventory=inventory,
        access_evidence=access_evidence,
    )

    if participant_audit is not None:
        audit_sheet = workbook.create_sheet(f"{project_key}_Participant_Audit")
        _write_dataframe_sheet(
            audit_sheet,
            participant_audit,
            table_name=f"{project_key}ParticipantAudit",
            freeze="C2",
        )
        _style_audit_decisions(audit_sheet, participant_audit)

    if combined_milestone_summary is not None:
        combined_sheet = workbook.create_sheet("Combined_Milestone_Summary")
        _write_dataframe_sheet(
            combined_sheet,
            combined_milestone_summary,
            table_name=f"{project_key}CombinedSummary",
        )
        if "Actual/Target Ratio" in combined_milestone_summary.columns:
            ratio_column = (
                list(combined_milestone_summary.columns).index(
                    "Actual/Target Ratio"
                )
                + 1
            )
            for row in range(2, len(combined_milestone_summary) + 2):
                combined_sheet.cell(row, ratio_column).number_format = "0%"

    if data_quality_issues is not None:
        quality_sheet = workbook.create_sheet("Data_Quality_Issues")
        _write_dataframe_sheet(
            quality_sheet,
            data_quality_issues,
            table_name=f"{project_key}DataQuality",
        )

    workbook.save(path)
    return path


def write_csv_package(
    *,
    output_dir: Path,
    project_audits: Mapping[str, pd.DataFrame],
    project_summaries: Mapping[str, pd.DataFrame],
    combined: pd.DataFrame,
    data_quality: pd.DataFrame,
) -> list[Path]:
    """Write the six requested logical outputs as formula-safe UTF-8 CSVs."""

    output_dir.mkdir(parents=True, exist_ok=True)
    datasets: dict[str, pd.DataFrame] = {
        "NANO_Participant_Audit": project_audits.get("NANO", pd.DataFrame()),
        "NICO_Participant_Audit": project_audits.get("NICO", pd.DataFrame()),
        "NANO_Milestone_Summary": project_summaries.get("NANO", pd.DataFrame()),
        "NICO_Milestone_Summary": project_summaries.get("NICO", pd.DataFrame()),
        "Combined_Milestone_Summary": combined,
        "Data_Quality_Issues": data_quality,
    }
    paths: list[Path] = []
    for name, frame in datasets.items():
        path = output_dir / f"{name}.csv"
        path.write_bytes(dataframe_to_csv_bytes(frame))
        paths.append(path)
    return paths


__all__ = [
    "write_csv_package",
    "write_project_workbook",
]
