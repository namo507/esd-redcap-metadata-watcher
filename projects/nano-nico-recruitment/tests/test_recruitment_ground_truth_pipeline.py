from __future__ import annotations

import datetime as dt
from dataclasses import replace
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import pandas as pd
import pytest

from recruitment_config import (
    CATEGORIES,
    PROJECT_CONFIG,
    PROJECT_ORDER,
    DiagnosticDateField,
    FieldExpectation,
    ProjectConfig,
)
from recruitment_ground_truth import (
    AUDIT_REQUIRED_COLUMNS,
    DQ_COLUMNS,
    RecruitmentGroundTruthError,
    actuals_from_audit,
    collapse_participants,
    combined_summary,
    minimal_record_fields,
    normalize_metadata,
    summary_long,
    validate_metadata,
    validate_project_identity,
)
from recruitment_workbooks import write_csv_package, write_project_workbook


REPORT_DATE = dt.date(2026, 7, 24)


def _milestones(config: ProjectConfig) -> list[dt.date]:
    return [
        dt.date(year, month, 1)
        for year in range(
            config.milestone_start.year, config.milestone_end.year + 1
        )
        for month in config.milestone_months
        if config.milestone_start
        <= dt.date(year, month, 1)
        <= config.milestone_end
    ]


def _choice_text(expectation: FieldExpectation) -> str:
    return " | ".join(
        f"{code}, Synthetic choice {code}" for code in expectation.allowed_codes
    )


def _metadata_for(config: ProjectConfig) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for expectation in config.field_expectations:
        rows.append(
            {
                "field_name": expectation.field_name,
                "form_name": expectation.form_name,
                "field_type": expectation.field_type,
                "select_choices_or_calculations": _choice_text(expectation),
                "text_validation_type_or_show_slider_number": "",
            }
        )
    existing = {row["field_name"] for row in rows}
    for diagnostic in config.diagnostic_date_fields:
        if diagnostic.field_name not in existing:
            rows.append(
                {
                    "field_name": diagnostic.field_name,
                    "form_name": "synthetic_diagnostic_dates",
                    "field_type": "text",
                    "select_choices_or_calculations": "",
                    "text_validation_type_or_show_slider_number": "date_ymd",
                }
            )
    return pd.DataFrame(rows)


def _append_metadata_field(
    metadata: pd.DataFrame,
    *,
    field_name: str,
    form_name: str,
    field_type: str,
    choices: str = "",
    validation: str = "",
) -> pd.DataFrame:
    return pd.concat(
        [
            metadata,
            pd.DataFrame(
                [
                    {
                        "field_name": field_name,
                        "form_name": form_name,
                        "field_type": field_type,
                        "select_choices_or_calculations": choices,
                        "text_validation_type_or_show_slider_number": validation,
                    }
                ]
            ),
        ],
        ignore_index=True,
    )


@pytest.mark.parametrize("project_key", PROJECT_ORDER)
def test_project_config_keeps_verified_identity_and_mappings_separate(
    project_key: str,
) -> None:
    config = PROJECT_CONFIG[project_key]

    assert config.expected_project_id is not None
    assert config.expected_project_title
    assert config.record_id != config.race_field != config.eth_field
    assert config.enrollment_status_field is None
    assert config.date_anchor is None
    assert config.record_presence_confirms_enrollment is False
    assert all(
        diagnostic.field_name != config.date_anchor
        for diagnostic in config.diagnostic_date_fields
    )

    if project_key == "NANO":
        assert config.race_field == "fif_childrace"
        assert config.eth_secondary == ()
    else:
        assert config.race_field == "race"
        assert config.eth_secondary == (("ethnicity", (0,)),)
        assert config.dual_field == "dual_enrolled"


@pytest.mark.parametrize("project_key", PROJECT_ORDER)
def test_validate_project_identity_accepts_only_the_configured_project(
    project_key: str,
) -> None:
    config = PROJECT_CONFIG[project_key]
    valid = {
        "project_id": config.expected_project_id,
        "project_title": config.expected_project_title,
    }
    validate_project_identity(project_key, valid, config)

    with pytest.raises(RecruitmentGroundTruthError, match="expected"):
        validate_project_identity(
            project_key,
            {
                "project_id": 999999,
                "project_title": config.expected_project_title,
            },
            config,
        )
    with pytest.raises(RecruitmentGroundTruthError, match="project title"):
        validate_project_identity(
            project_key,
            {
                "project_id": config.expected_project_id,
                "project_title": "Synthetic Wrong Project",
            },
            config,
        )


@pytest.mark.parametrize("project_key", PROJECT_ORDER)
def test_validate_metadata_accepts_exact_schema_and_index_field_name(
    project_key: str,
) -> None:
    config = PROJECT_CONFIG[project_key]
    indexed_metadata = _metadata_for(config).set_index("field_name")

    normalized, checks = validate_metadata(
        project_key, indexed_metadata, config
    )

    assert set(normalized["field_name"]) >= {
        expectation.field_name for expectation in config.field_expectations
    }
    assert len(checks) == len(config.field_expectations)
    assert checks["result"].eq("PASS").all()


@pytest.mark.parametrize("drift", ["missing_field", "wrong_form", "missing_code"])
def test_validate_metadata_fails_closed_on_schema_drift(drift: str) -> None:
    config = PROJECT_CONFIG["NICO"]
    metadata = _metadata_for(config)

    if drift == "missing_field":
        metadata = metadata.loc[metadata["field_name"].ne(config.race_field)]
        expected_field = config.race_field
    elif drift == "wrong_form":
        metadata.loc[
            metadata["field_name"].eq(config.eth_field), "form_name"
        ] = "synthetic_wrong_form"
        expected_field = config.eth_field
    else:
        metadata.loc[
            metadata["field_name"].eq(config.race_field),
            "select_choices_or_calculations",
        ] = "1, Synthetic choice 1"
        expected_field = config.race_field

    with pytest.raises(RecruitmentGroundTruthError, match=expected_field):
        validate_metadata("NICO", metadata, config)


def test_normalize_metadata_rejects_empty_blank_or_missing_field_names() -> None:
    with pytest.raises(RecruitmentGroundTruthError, match="empty"):
        normalize_metadata(pd.DataFrame())
    with pytest.raises(RecruitmentGroundTruthError, match="field_name"):
        normalize_metadata(pd.DataFrame({"unrelated": ["value"]}))
    with pytest.raises(RecruitmentGroundTruthError, match="blank field"):
        normalize_metadata(pd.DataFrame({"field_name": [""]}))


@pytest.mark.parametrize(
    ("project_key", "expected_fields"),
    [
        (
            "NANO",
            [
                "demo_id",
                "fif_childrace",
                "fif_childethnicity",
                "demo_ineligible",
                "demo_unenrolled",
                "demo_exclude",
            ],
        ),
        (
            "NICO",
            [
                "id",
                "race",
                "fif_childethnicity",
                "ethnicity",
                "demo_ineligible",
                "demo_unenrolled",
                "demo_exclude",
                "dual_enrolled",
            ],
        ),
    ],
)
def test_minimal_record_fields_are_exact_ordered_and_source_backed(
    project_key: str, expected_fields: list[str]
) -> None:
    config = PROJECT_CONFIG[project_key]
    metadata = _metadata_for(config)

    assert minimal_record_fields(config, metadata) == expected_fields
    assert not {
        diagnostic.field_name for diagnostic in config.diagnostic_date_fields
    }.intersection(expected_fields)

    without_last_required_field = metadata.loc[
        metadata["field_name"].ne(expected_fields[-1])
    ]
    result = minimal_record_fields(config, without_last_required_field)
    assert expected_fields[-1] not in result
    assert len(result) == len(set(result))


def _nico_records() -> pd.DataFrame:
    config = PROJECT_CONFIG["NICO"]
    race_columns = {
        f"{config.race_field}___{code}": 0
        for code in (
            *config.race_minority_codes,
            *config.race_white_codes,
            *config.race_unknown_codes,
            *config.race_hispanic_codes,
        )
    }

    def row(
        participant_id: str,
        event: str,
        *,
        race_codes: tuple[int, ...] = (),
        primary_ethnicity: Any = pd.NA,
        secondary_ethnicity: Any = pd.NA,
        ineligible: Any = 0,
        unenrolled: Any = 0,
        review: Any = 0,
        dual: Any = 0,
        visit_date: Any = pd.NA,
    ) -> dict[str, Any]:
        output: dict[str, Any] = {
            "id": participant_id,
            "redcap_event_name": event,
            **race_columns,
            "fif_childethnicity": primary_ethnicity,
            "ethnicity": secondary_ethnicity,
            "demo_ineligible": ineligible,
            "demo_unenrolled": unenrolled,
            "demo_exclude___1": review,
            "dual_enrolled": dual,
            "visit_date": visit_date,
            "dob": pd.NA,
        }
        for code in race_codes:
            output[f"race___{code}"] = 1
        return output

    return pd.DataFrame(
        [
            row(
                "FAKE-NICO-001",
                "baseline_arm_1",
                race_codes=(1,),
                secondary_ethnicity=0,
                dual=0,
                visit_date="2026-01-15",
            ),
            row(
                "FAKE-NICO-001",
                "followup_arm_1",
                race_codes=(6,),
                dual=0,
            ),
            row(
                "FAKE-NICO-002",
                "baseline_arm_1",
                race_codes=(7,),
                primary_ethnicity=3,
                secondary_ethnicity=2,
            ),
            row(
                "FAKE-NICO-003",
                "baseline_arm_1",
                race_codes=(6,),
                primary_ethnicity=1,
                secondary_ethnicity=1,
            ),
            row(
                "FAKE-NICO-004",
                "baseline_arm_1",
                primary_ethnicity=2,
                ineligible=0,
            ),
            row(
                "FAKE-NICO-004",
                "followup_arm_1",
                ineligible=1,
            ),
            row(
                "FAKE-NICO-005",
                "baseline_arm_1",
                race_codes=(2,),
                primary_ethnicity=2,
                review=1,
            ),
            row(
                "FAKE-NICO-006",
                "baseline_arm_1",
                race_codes=(6,),
                primary_ethnicity=2,
                unenrolled=1,
            ),
        ]
    )


@pytest.fixture
def nico_collapsed() -> tuple[pd.DataFrame, pd.DataFrame]:
    config = PROJECT_CONFIG["NICO"]
    return collapse_participants(
        project_key="NICO",
        records=_nico_records(),
        metadata=_metadata_for(config),
        config=config,
        milestones=_milestones(config),
        report_date=REPORT_DATE,
    )


def test_collapse_across_events_preserves_zero_code_and_multirace(
    nico_collapsed: tuple[pd.DataFrame, pd.DataFrame],
) -> None:
    audit, issues = nico_collapsed
    indexed = audit.set_index("participant_id")
    participant = indexed.loc["FAKE-NICO-001"]

    assert len(audit) == 6
    assert audit["participant_id"].is_unique
    assert participant["source_row_count"] == 2
    assert participant["race_selection_count"] == 2
    assert participant["racial_minority_flag"] == True  # noqa: E712
    assert participant["hispanic_ethnicity_flag"] == True  # noqa: E712
    assert "ethnicity=0" in participant["raw_ethnicity"]
    assert participant["raw_dual_enrollment"] == "0"
    assert participant["candidate_visit_date_first"] == "2026-01-15"
    assert participant["milestone_period"] == ""
    assert (
        issues.loc[
            issues["participant_id"].eq("FAKE-NICO-001"), "issue_type"
        ]
        .eq("multiple_race_selections")
        .any()
    )


def test_unknown_and_conflicting_ethnicity_remain_auditable(
    nico_collapsed: tuple[pd.DataFrame, pd.DataFrame],
) -> None:
    audit, issues = nico_collapsed
    indexed = audit.set_index("participant_id")

    unknown = indexed.loc["FAKE-NICO-002"]
    assert unknown["racial_minority_flag"] == False  # noqa: E712
    assert pd.isna(unknown["hispanic_ethnicity_flag"])
    unknown_issues = set(
        issues.loc[
            issues["participant_id"].eq("FAKE-NICO-002"), "issue_type"
        ]
    )
    assert {"unknown_other_race", "unknown_declined_ethnicity"} <= unknown_issues

    conflict = indexed.loc["FAKE-NICO-003"]
    assert conflict["hispanic_ethnicity_flag"] == True  # noqa: E712
    assert (
        issues.loc[
            issues["participant_id"].eq("FAKE-NICO-003"), "issue_type"
        ]
        .eq("conflicting_ethnicity_sources")
        .any()
    )


def test_exclusion_and_review_flags_are_collapsed_and_explained(
    nico_collapsed: tuple[pd.DataFrame, pd.DataFrame],
) -> None:
    audit, issues = nico_collapsed
    indexed = audit.set_index("participant_id")

    ineligible = indexed.loc["FAKE-NICO-004"]
    assert ineligible["source_row_count"] == 2
    assert ineligible["included_in_recruitment_count"] == False  # noqa: E712
    assert ineligible["inclusion_decision"] == "excluded"
    assert ineligible["exclusion_reason"] == "demo_ineligible=1"

    unenrolled = indexed.loc["FAKE-NICO-006"]
    assert unenrolled["included_in_recruitment_count"] == False  # noqa: E712
    assert unenrolled["exclusion_reason"] == "demo_unenrolled=1"

    review = indexed.loc["FAKE-NICO-005"]
    assert review["review_flags"] == "demo_exclude=1"
    assert review["inclusion_decision"] == "unresolved"

    issue_pairs = set(
        issues.loc[
            issues["participant_id"].isin(
                ["FAKE-NICO-004", "FAKE-NICO-005", "FAKE-NICO-006"]
            ),
            ["participant_id", "issue_type"],
        ].itertuples(index=False, name=None)
    )
    assert ("FAKE-NICO-004", "excluded_status") in issue_pairs
    assert ("FAKE-NICO-005", "review_status") in issue_pairs
    assert ("FAKE-NICO-006", "excluded_status") in issue_pairs


def test_hard_exclusion_field_is_requested_and_takes_precedence() -> None:
    base = PROJECT_CONFIG["NICO"]
    config = replace(base, hard_exclude_flags=("synthetic_admin_record",))
    metadata = _append_metadata_field(
        _metadata_for(config),
        field_name="synthetic_admin_record",
        form_name="synthetic_admin",
        field_type="yesno",
        choices="0, No | 1, Yes",
    )
    records = _nico_records().iloc[[0]].copy()
    records["synthetic_admin_record"] = 1

    assert "synthetic_admin_record" in minimal_record_fields(config, metadata)
    audit, issues = collapse_participants(
        project_key="NICO",
        records=records,
        metadata=metadata,
        config=config,
        milestones=_milestones(config),
        report_date=REPORT_DATE,
    )

    assert audit.loc[0, "included_in_recruitment_count"] == False  # noqa: E712
    assert audit.loc[0, "exclusion_reason"] == "synthetic_admin_record=1"
    assert issues["issue_type"].eq("excluded_status").any()


def test_unresolved_enrollment_mapping_does_not_use_diagnostic_dates(
    nico_collapsed: tuple[pd.DataFrame, pd.DataFrame],
) -> None:
    audit, issues = nico_collapsed
    nonexcluded = audit.loc[
        ~audit["participant_id"].isin(["FAKE-NICO-004", "FAKE-NICO-006"])
    ]

    assert nonexcluded["included_in_recruitment_count"].isna().all()
    assert nonexcluded["milestone_period"].eq("").all()
    assert (
        audit.set_index("participant_id").loc[
            "FAKE-NICO-001", "candidate_visit_date_first"
        ]
        == "2026-01-15"
    )
    project_issues = set(
        issues.loc[issues["participant_id"].eq(""), "issue_type"]
    )
    assert {
        "unresolved_enrollment_status_mapping",
        "unresolved_enrollment_date_mapping",
    } <= project_issues

    actuals = actuals_from_audit(
        audit, PROJECT_CONFIG["NICO"], _milestones(PROJECT_CONFIG["NICO"])
    )
    assert all(
        value is None
        for category_values in actuals.values()
        for value in category_values
    )


def test_unconfirmed_mapping_preserves_published_history_only() -> None:
    empty_audit = pd.DataFrame(columns=AUDIT_REQUIRED_COLUMNS)
    nano = PROJECT_CONFIG["NANO"]
    nico = PROJECT_CONFIG["NICO"]

    nano_actuals = actuals_from_audit(empty_audit, nano, _milestones(nano))
    for category in CATEGORIES:
        expected_seed = nano.historical_actuals[category]
        assert nano_actuals[category][: len(expected_seed)] == expected_seed
        assert all(
            value is None
            for value in nano_actuals[category][len(expected_seed) :]
        )

    nico_actuals = actuals_from_audit(empty_audit, nico, _milestones(nico))
    assert all(
        value is None
        for values in nico_actuals.values()
        for value in values
    )


def test_combined_summary_never_treats_one_project_as_the_total() -> None:
    summaries: dict[str, pd.DataFrame] = {}
    for project_key in PROJECT_ORDER:
        config = PROJECT_CONFIG[project_key]
        milestones = _milestones(config)
        summaries[project_key] = summary_long(
            project_key=project_key,
            config=config,
            milestones=milestones,
            actuals=actuals_from_audit(
                pd.DataFrame(columns=AUDIT_REQUIRED_COLUMNS),
                config,
                milestones,
            ),
        )

    combined = combined_summary(summaries)
    combined_rows = combined.loc[combined["Project"].eq("COMBINED")]

    assert len(combined_rows) == len(_milestones(PROJECT_CONFIG["NANO"])) * 3
    # NANO has published values, but NICO does not. The combined value must
    # remain unavailable rather than silently becoming the NANO-only value.
    assert combined_rows["Actual"].isna().all()
    assert combined_rows["Previous Target"].isna().all()
    assert combined_rows["Current Target"].isna().all()
    assert combined_rows["Status"].eq("N/A").all()


def _confirmed_config_and_metadata() -> tuple[ProjectConfig, pd.DataFrame]:
    base = PROJECT_CONFIG["NANO"]
    status_expectation = FieldExpectation(
        "synthetic_enrollment_status",
        "synthetic_enrollment",
        "radio",
        ("0", "1"),
    )
    date_expectation = FieldExpectation(
        "synthetic_enrollment_date",
        "synthetic_enrollment",
        "text",
    )
    config = replace(
        base,
        enrollment_status_field=status_expectation.field_name,
        enrolled_status_codes=("1",),
        date_anchor=date_expectation.field_name,
        record_presence_confirms_enrollment=False,
        diagnostic_date_fields=(
            DiagnosticDateField("visit_date", "Synthetic visit date"),
        ),
        field_expectations=(
            *base.field_expectations,
            status_expectation,
            date_expectation,
        ),
        historical_actuals={},
    )
    metadata = _metadata_for(config)
    metadata.loc[
        metadata["field_name"].eq("synthetic_enrollment_date"),
        "text_validation_type_or_show_slider_number",
    ] = "date_ymd"
    return config, metadata


def _confirmed_records() -> pd.DataFrame:
    race_defaults = {
        f"fif_childrace___{code}": 0 for code in range(1, 7)
    }

    def row(
        participant_id: str,
        status: Any,
        enrollment_date: Any,
        *,
        race_code: int = 5,
        ethnicity: Any = 2,
        event: str = "baseline_arm_1",
    ) -> dict[str, Any]:
        output: dict[str, Any] = {
            "demo_id": participant_id,
            "redcap_event_name": event,
            **race_defaults,
            "fif_childethnicity": ethnicity,
            "demo_ineligible": 0,
            "demo_unenrolled": 0,
            "demo_exclude___1": 0,
            "synthetic_enrollment_status": status,
            "synthetic_enrollment_date": enrollment_date,
            "visit_date": "2025-01-01",
        }
        output[f"fif_childrace___{race_code}"] = 1
        return output

    return pd.DataFrame(
        [
            row("FAKE-DATE-001", 1, "2026-04-01", race_code=1, ethnicity=1),
            row("FAKE-DATE-002", 1, "2026-04-02"),
            row("FAKE-DATE-003", 1, "2026-08-01", race_code=2),
            row("FAKE-DATE-004", 1, "2026-08-02", ethnicity=1),
            row("FAKE-DATE-005", 0, "2026-01-01"),
            row("FAKE-DATE-006", 1, pd.NA),
            row("FAKE-DATE-007", 1, "not-a-date"),
            row("FAKE-DATE-008", 1, "2026-04-01", event="baseline_arm_1"),
            row("FAKE-DATE-008", 1, "2026-04-02", event="followup_arm_1"),
        ]
    )


def test_confirmed_enrollment_date_boundaries_use_replaced_config() -> None:
    config, metadata = _confirmed_config_and_metadata()
    assert PROJECT_CONFIG["NANO"].date_anchor is None
    assert config.date_anchor == "synthetic_enrollment_date"

    milestones = [
        dt.date(2026, 4, 1),
        dt.date(2026, 8, 1),
        dt.date(2026, 12, 1),
    ]
    audit, issues = collapse_participants(
        project_key="NANO",
        records=_confirmed_records(),
        metadata=metadata,
        config=config,
        milestones=milestones,
        report_date=dt.date(2026, 8, 1),
    )
    indexed = audit.set_index("participant_id")

    assert indexed.loc["FAKE-DATE-001", "milestone_period"] == "2026-04-01"
    assert indexed.loc["FAKE-DATE-002", "milestone_period"] == "2026-08-01"
    assert indexed.loc["FAKE-DATE-003", "milestone_period"] == "2026-08-01"
    assert (
        indexed.loc["FAKE-DATE-003", "included_in_recruitment_count"]
        == True  # noqa: E712
    )
    assert (
        indexed.loc["FAKE-DATE-004", "included_in_recruitment_count"]
        == False  # noqa: E712
    )
    assert (
        indexed.loc["FAKE-DATE-004", "exclusion_reason"]
        == "ENROLLMENT_AFTER_REPORT_CUTOFF"
    )
    assert (
        indexed.loc["FAKE-DATE-005", "exclusion_reason"]
        == "NOT_CONFIRMED_ENROLLED"
    )

    actuals = actuals_from_audit(audit, config, milestones)
    assert actuals == {
        "Total": [1, 3, 3],
        "Minority": [1, 2, 2],
        "Hispanic": [1, 1, 1],
    }

    issue_types = set(issues["issue_type"])
    assert {
        "missing_enrollment_date",
        "invalid_enrollment_date",
        "conflicting_enrollment_dates",
        "unassignable_milestone",
    } <= issue_types


def _synthetic_audit(project_key: str, participant_id: str) -> pd.DataFrame:
    row = {column: "" for column in AUDIT_REQUIRED_COLUMNS}
    row.update(
        {
            "participant_id": participant_id,
            "project": project_key,
            "included_in_recruitment_count": pd.NA,
            "racial_minority_flag": True,
            "hispanic_ethnicity_flag": False,
            "inclusion_decision": "unresolved",
            "source_row_count": 1,
        }
    )
    return pd.DataFrame([row])


def _synthetic_quality(project_key: str) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "project": project_key,
                "participant_id": "FAKE-QUALITY-001",
                "issue_type": "synthetic_issue",
                "field_name": "synthetic_field",
                "raw_value": "synthetic",
                "detail": "Synthetic test-only issue.",
            }
        ],
        columns=DQ_COLUMNS,
    )


def test_exact_two_restricted_workbooks_have_expected_sheets_and_formulas(
    tmp_path: Path,
) -> None:
    restricted_dir = tmp_path / "restricted"
    summaries: dict[str, pd.DataFrame] = {}
    actuals_by_project: dict[str, dict[str, list[int | None]]] = {}
    for project_key in PROJECT_ORDER:
        config = PROJECT_CONFIG[project_key]
        milestones = _milestones(config)
        actuals = actuals_from_audit(
            pd.DataFrame(columns=AUDIT_REQUIRED_COLUMNS),
            config,
            milestones,
        )
        actuals_by_project[project_key] = actuals
        summaries[project_key] = summary_long(
            project_key=project_key,
            config=config,
            milestones=milestones,
            actuals=actuals,
        )
    combined = combined_summary(summaries)

    for project_key in PROJECT_ORDER:
        config = PROJECT_CONFIG[project_key]
        write_project_workbook(
            path=restricted_dir
            / f"{project_key.lower()}_recruitment_ground_truth_2026-07-24.xlsx",
            project_key=project_key,
            config=config,
            milestones=_milestones(config),
            actuals=actuals_by_project[project_key],
            report_date=REPORT_DATE,
            project_info={
                "project_id": config.expected_project_id,
                "project_title": config.expected_project_title,
            },
            inventory={
                "participant_records": 1,
                "unresolved_inclusion": 1,
                "definitely_excluded": 0,
            },
            access_evidence={
                "project_info": True,
                "metadata": True,
                "minimal_records": True,
            },
            participant_audit=_synthetic_audit(
                project_key, f"FAKE-{project_key}-WORKBOOK-001"
            ),
            combined_milestone_summary=combined,
            data_quality_issues=_synthetic_quality(project_key),
        )

    expected_names = [
        "nano_recruitment_ground_truth_2026-07-24.xlsx",
        "nico_recruitment_ground_truth_2026-07-24.xlsx",
    ]
    assert sorted(path.name for path in restricted_dir.iterdir()) == expected_names

    for project_key in PROJECT_ORDER:
        path = (
            restricted_dir
            / f"{project_key.lower()}_recruitment_ground_truth_2026-07-24.xlsx"
        )
        workbook = load_workbook(path, data_only=False)
        assert workbook.sheetnames == [
            f"{project_key}_Milestone_Summary",
            f"{project_key}_Participant_Audit",
            "Combined_Milestone_Summary",
            "Data_Quality_Issues",
        ]
        assert workbook.properties.subject == "Restricted participant audit"

        summary = workbook[f"{project_key}_Milestone_Summary"]
        formulas = [
            cell.value
            for row in summary.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        assert len(formulas) == len(_milestones(PROJECT_CONFIG[project_key])) * 6
        assert (
            summary["B9"].value
            == '=IF(OR(B8="",B7="",B7=0),"N/A",B8/B7)'
        )
        assert (
            summary["B10"].value
            == '=IF(OR(B7="",B7=0),"N/A",IF(B8="","Not reported",IF(B8>=B7,"On target","Behind")))'
        )
        assert summary.freeze_panes == "B5"
        assert summary.sheet_view.showGridLines is False
        assert summary.max_column == 15
        workbook.close()


def test_csv_package_has_exact_six_names_and_formula_safe_fake_data(
    tmp_path: Path,
) -> None:
    nano_audit = _synthetic_audit("NANO", "=FAKE-NANO-CSV-001")
    nico_audit = _synthetic_audit("NICO", "FAKE-NICO-CSV-001")
    summaries = {
        project_key: summary_long(
            project_key=project_key,
            config=PROJECT_CONFIG[project_key],
            milestones=_milestones(PROJECT_CONFIG[project_key]),
            actuals=actuals_from_audit(
                pd.DataFrame(columns=AUDIT_REQUIRED_COLUMNS),
                PROJECT_CONFIG[project_key],
                _milestones(PROJECT_CONFIG[project_key]),
            ),
        )
        for project_key in PROJECT_ORDER
    }
    combined = combined_summary(summaries)
    quality = pd.concat(
        [_synthetic_quality("NANO"), _synthetic_quality("NICO")],
        ignore_index=True,
    )
    output_dir = tmp_path / "csv_package"

    paths = write_csv_package(
        output_dir=output_dir,
        project_audits={"NANO": nano_audit, "NICO": nico_audit},
        project_summaries=summaries,
        combined=combined,
        data_quality=quality,
    )

    expected_names = [
        "NANO_Participant_Audit.csv",
        "NICO_Participant_Audit.csv",
        "NANO_Milestone_Summary.csv",
        "NICO_Milestone_Summary.csv",
        "Combined_Milestone_Summary.csv",
        "Data_Quality_Issues.csv",
    ]
    assert [path.name for path in paths] == expected_names
    assert sorted(path.name for path in output_dir.iterdir()) == sorted(
        expected_names
    )
    for path in paths:
        assert path.read_bytes().startswith(b"\xef\xbb\xbf")
        assert path.read_text(encoding="utf-8-sig").strip()
    assert "'=FAKE-NANO-CSV-001" in (
        output_dir / "NANO_Participant_Audit.csv"
    ).read_text(encoding="utf-8-sig")
