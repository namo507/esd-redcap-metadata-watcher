"""Stakeholder-friendly bot analysis helpers for the caregiver study.

This module provides simple, presentation-ready visualisations, plain-English
column names, and a multi-sheet Excel export targeted at non-technical reviewers
(e.g., Dr. Bradshaw).  It reads the pipeline's committed outputs and REDCap
caches without duplicating computation logic.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
import yaml
from openpyxl.utils import get_column_letter


# ── Friendly column names ───────────────────────────────────────────────────

RULE_FRIENDLY_NAMES: dict[str, str] = {
    "rule_R1": "Survey Too Fast (<11.57 min)",
    "rule_R2": "Attitudes Too Fast (<7.85 min)",
    "rule_R3": "Any Section Below Speed Floor",
    "rule_R4": "Flat-Line Responses (Low Variation)",
    "rule_R5": "Duplicate Response Pattern",
    "rule_R6": "Bursty Submission Timing",
    "rule_R7": "Near-Duplicate Open Text",
    "rule_R8": "Illogical Family Info",
    "rule_R9": "Impossible Demographics",
    "rule_R10": "Attention Check Failed",
}

COLUMN_FRIENDLY_NAMES: dict[str, str] = {
    "record_id": "Record ID",
    "tier_label": "Trust Category",
    "tier": "Tier",
    "giftcard_decision": "Payment Decision",
    "decision_reason": "Reason",
    "tier1_hit_count": "Hard Check Violations",
    "tier1_rule_combo": "Hard Rules Triggered",
    "soft_triggered_rules": "Soft Rules Triggered",
    "classification": "Classification",
    "action": "Recommended Action",
    "confirmed_reason": "Reason for Rejection",
    "total_rules_fired": "Total Rules Fired",
    "tier_1_rules_fired": "Hard Rules Fired",
    "soft_rules_fired": "Soft Rules Fired",
    **RULE_FRIENDLY_NAMES,
}

# ── Category colour palette (colourblind-friendly) ──────────────────────────

CATEGORY_COLORS: dict[str, str] = {
    "Confirmed Bot": "#A85D75",       # pink
    "Needs Human Review": "#C67C2D",  # orange
    "Likely Real Caregivers": "#B08A2E",  # gold
    "Real Caregivers": "#1F5A7A",     # blue
}

CATEGORY_ORDER = [
    "Confirmed Bot",
    "Needs Human Review",
    "Likely Real Caregivers",
    "Real Caregivers",
]


PROJECT_LABELS: dict[str, str] = {
    "clean_4797": "Study 1 - Verified (4797)",
    "dirty_4581": "Study 2 - Legacy (4581)",
}

PAYMENT_DECISION_LABELS: dict[str, str] = {
    "auto_eligible": "Cleared for payment now",
    "eligible_low_risk_review": "Low-risk provisional approval",
    "manual_review": "Needs human review",
    "do_not_pay_pending_adjudication": "Confirmed bot / reject",
}

SUBCLASSIFICATION_LABELS: dict[str, str] = {
    "Real Caregiver": "No rule violations",
    "Likely Real Caregiver": "One soft check only",
    "Needs Review — Single Hard Flag": "Single hard-check flag",
    "Needs Review — Multiple Hard Flags": "Multiple hard-check flags",
    "Needs Review — Multiple Soft Flags": "Two or more soft-check flags",
    "Confirmed Bot": "Hard logic contradiction with speed flags",
}

GENDER_LABELS: dict[str, str] = {
    "0": "Prefer not to answer",
    "1": "Man",
    "2": "Woman",
    "3": "Non-binary",
    "4": "Other",
}

RACE_CHECKBOX_LABELS: dict[str, str] = {
    "demo_maternalrace___1": "American Indian/Alaska Native",
    "demo_maternalrace___2": "Asian",
    "demo_maternalrace___3": "Native Hawaiian or Other Pacific Islander",
    "demo_maternalrace___4": "Black or African American",
    "demo_maternalrace___5": "White",
    "demo_maternalrace___6": "Unknown",
    "demo_maternalrace___7": "Other",
}

ETHNICITY_CHECKBOX_LABELS: dict[str, str] = {
    "demo_maternalethnicity___1": "Hispanic/Latino",
    "demo_maternalethnicity___2": "Not Hispanic/Latino",
    "demo_maternalethnicity___3": "Unknown",
    "demo_maternalethnicity___4": "Other",
}

WORKBOOK_EXPORT_COLUMNS = [
    "Study",
    "Record ID",
    "Stakeholder Bucket",
    "Payment Decision",
    "Trust Tier",
    "Reason for Bucket",
    "Hard Rules Triggered",
    "Soft Rules Triggered",
    "All Rules Triggered",
    "Hard Check Violations",
    "Soft Check Violations",
    "Total Rules Fired",
    "Total Survey Time (min)",
    "Attitudes Section Time (min)",
    "Extreme Fast (R1 + R2)",
    "Duplicate Response Pattern",
    "Bursty Submission Timing",
    "Branching / Family Logic Issue",
    "Impossible Demographics",
    "Caregiver Age",
    "Gender",
    "Race",
    "Ethnicity",
    "Status-Quo Cluster",
]


# ── Data loading ────────────────────────────────────────────────────────────

def load_config(project_dir: Path) -> dict:
    with (project_dir / "config.yaml").open("r", encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def load_master_summary(output_dir: Path) -> pd.DataFrame:
    """Load the pre-computed master summary with 4 categories."""
    df = pd.read_csv(output_dir / "table_43_master_summary.csv")
    return df


def load_detailed_breakdown(output_dir: Path) -> pd.DataFrame:
    """Load the detailed sub-category breakdown."""
    df = pd.read_csv(output_dir / "table_44_detailed_subcategory_breakdown.csv")
    return df


def load_confirmed_bots(output_dir: Path) -> pd.DataFrame:
    """Load confirmed bot records with friendly column names."""
    df = pd.read_csv(output_dir / "table_45_confirmed_bot_records.csv")
    return df.rename(columns=COLUMN_FRIENDLY_NAMES)


def load_record_flags(output_dir: Path) -> pd.DataFrame:
    return pd.read_parquet(output_dir / "record_flags.parquet")


def load_rule_definitions(output_dir: Path) -> pd.DataFrame:
    df = pd.read_csv(output_dir / "table_14_fraud_rule_definitions.csv")
    df["Friendly Name"] = df["rule"].map(
        {k.replace("rule_", ""): v for k, v in RULE_FRIENDLY_NAMES.items()}
    )
    return df


def load_dirty_records(cache_dir: Path) -> pd.DataFrame:
    """Load the dirty 4581 REDCap cache (latest parquet)."""
    return _load_project_records(cache_dir, 4581, "dirty_4581")


def load_record_classification(output_dir: Path) -> pd.DataFrame:
    """Load the row-level stakeholder classification table."""
    df = pd.read_csv(output_dir / "table_41_distinct_record_classification.csv")
    df["record_id"] = df["record_id"].astype(str)
    return df


def _load_project_records(cache_dir: Path, project_id: int, source_project: str) -> pd.DataFrame:
    matches = sorted(cache_dir.glob(f"{project_id}_record_*.parquet"))
    if not matches:
        raise FileNotFoundError(f"No {project_id} record cache found.")
    df = pd.read_parquet(matches[-1]).copy()
    df["record_id"] = df["record_id"].astype(str)
    df["source_project"] = source_project
    return df


def load_combined_records(cache_dir: Path) -> pd.DataFrame:
    """Load the latest REDCap record caches for both studies."""
    return pd.concat(
        [
            _load_project_records(cache_dir, 4797, "clean_4797"),
            _load_project_records(cache_dir, 4581, "dirty_4581"),
        ],
        ignore_index=True,
        sort=False,
    )


# ── Classification helper ───────────────────────────────────────────────────

def classify_records(
    record_flags: pd.DataFrame,
    confirmed_bot_ids: set[str],
) -> pd.DataFrame:
    """Assign each record to one of 4 stakeholder categories.

    Categories:
        - Confirmed Bot (6 identified records)
        - Needs Human Review (Tier 1/2 + ≥2 soft flags)
        - Likely Real Caregivers (Tier 3, one soft flag only)
        - Real Caregivers (Tier 4, no flags)
    """
    df = record_flags.copy()
    df["record_id_str"] = df["record_id"].astype(str)

    conditions = [
        df["record_id_str"].isin(confirmed_bot_ids),
        df["tier"].isin([1, 2]),
        df["tier"] == 3,
        df["tier"] == 4,
    ]
    choices = [
        "Confirmed Bot",
        "Needs Human Review",
        "Likely Real Caregivers",
        "Real Caregivers",
    ]
    df["Category"] = np.select(conditions, choices, default="Needs Human Review")
    df.drop(columns=["record_id_str"], inplace=True)
    return df


# ── Timing helpers ──────────────────────────────────────────────────────────

def compute_timing(
    dirty_records: pd.DataFrame,
    record_flags: pd.DataFrame,
) -> pd.DataFrame:
    """Merge timing data from REDCap cache into the classified flags."""
    time_fields = ["get_time_fif", "get_time_val", "get_time_tfa", "get_time_demo"]
    rec = dirty_records[["record_id"] + time_fields].copy()
    rec["record_id"] = rec["record_id"].astype(str)

    # Convert timing fields to numeric minutes
    for col in time_fields:
        rec[col] = pd.to_numeric(rec[col], errors="coerce")

    rec["Total Survey Time (min)"] = rec[time_fields].sum(axis=1)
    rec["Attitudes Section Time (min)"] = rec["get_time_tfa"]

    flags = record_flags.copy()
    flags["record_id"] = flags["record_id"].astype(str)

    merged = flags.merge(
        rec[["record_id", "Total Survey Time (min)", "Attitudes Section Time (min)"]],
        on="record_id",
        how="left",
    )
    return merged


def _normalise_rule_text(value: object) -> str:
    text = str(value).strip()
    if not text or text.lower() == "none" or text.lower() == "nan":
        return "None"
    return ", ".join(part.strip() for part in text.split(",") if part.strip())


def _yes_no(series: pd.Series) -> pd.Series:
    return np.where(series.fillna(False).astype(bool), "Yes", "No")


def _format_checkbox_selection(
    frame: pd.DataFrame,
    mapping: dict[str, str],
    other_flag: Optional[str] = None,
    other_text_col: Optional[str] = None,
) -> pd.Series:
    def build_value(row: pd.Series) -> str:
        selections: list[str] = []
        for column, label in mapping.items():
            if str(row.get(column, "")).strip() == "1":
                selections.append(label)
        if other_flag and other_text_col and str(row.get(other_flag, "")).strip() == "1":
            other_text = str(row.get(other_text_col, "")).strip()
            if other_text:
                selections = [
                    f"Other: {other_text}" if entry == "Other" else entry
                    for entry in selections
                ]
        return "; ".join(selections) if selections else "Not answered"

    return frame.apply(build_value, axis=1)


def _format_gender(series: pd.Series) -> pd.Series:
    formatted = series.astype("string").str.strip().map(GENDER_LABELS)
    return formatted.fillna("Not answered")


def build_stakeholder_record_view(
    output_dir: Path,
    cache_dir: Optional[Path] = None,
) -> pd.DataFrame:
    """Build a stakeholder-friendly record-level export with timing and demographics."""
    if cache_dir is None:
        cache_dir = output_dir.parent / "data_cache"

    classification = load_record_classification(output_dir)
    record_flags = load_record_flags(output_dir).copy()
    record_flags["record_id"] = record_flags["record_id"].astype(str)
    records = load_combined_records(cache_dir)

    merged = classification.merge(
        record_flags,
        on=["source_project", "project_id", "record_id", "tier", "tier_label"],
        how="left",
        validate="one_to_one",
    ).merge(
        records,
        on=["source_project", "record_id"],
        how="left",
        validate="one_to_one",
    )

    time_fields = ["get_time_fif", "get_time_val", "get_time_tfa", "get_time_demo"]
    for column in time_fields + ["age_check_demo"]:
        merged[column] = pd.to_numeric(merged[column], errors="coerce")

    merged["Total Survey Time (min)"] = merged[time_fields].sum(axis=1, min_count=1).round(2)
    merged["Attitudes Section Time (min)"] = merged["get_time_tfa"].round(2)
    merged["Caregiver Age"] = merged["age_check_demo"].where(
        merged["age_check_demo"].between(18, 100)
    ).round(1)
    merged["Gender"] = _format_gender(merged["demo_gender"])
    merged["Race"] = _format_checkbox_selection(
        merged,
        RACE_CHECKBOX_LABELS,
        other_flag="demo_maternalrace___7",
        other_text_col="demo_race_other",
    )
    merged["Ethnicity"] = _format_checkbox_selection(
        merged,
        ETHNICITY_CHECKBOX_LABELS,
        other_flag="demo_maternalethnicity___4",
        other_text_col="demo_ethnicity_other",
    )
    merged["Stakeholder Bucket"] = merged["classification"].map(
        {
            "Confirmed Bot": "Confirmed bot / reject",
            "Needs Review": "Needs human review",
            "Likely Real Caregiver": "Low-risk provisional approval",
            "Real Caregiver": "Cleared for payment now",
        }
    ).fillna(merged["classification"])
    merged["Record ID"] = merged["record_id"].astype(str)
    merged["Trust Tier"] = merged["tier_label"]
    merged["Payment Decision"] = merged["giftcard_decision"].map(PAYMENT_DECISION_LABELS).fillna(
        merged["giftcard_decision"]
    )
    merged["Reason for Bucket"] = merged["sub_classification"].map(SUBCLASSIFICATION_LABELS).fillna(
        merged["sub_classification"]
    )
    merged["Hard Rules Triggered"] = merged["tier1_triggered"].map(_normalise_rule_text)
    merged["Soft Rules Triggered"] = merged["soft_triggered"].map(_normalise_rule_text)
    merged["All Rules Triggered"] = merged["all_rules_triggered"].map(_normalise_rule_text)
    merged["Hard Check Violations"] = merged["tier1_hit_count"].fillna(0).astype(int)
    merged["Soft Check Violations"] = merged["suspicion_rule_count"].fillna(0).astype(int)
    merged["Total Rules Fired"] = merged["total_rules_fired"].fillna(0).astype(int)
    merged["Duplicate Response Pattern"] = _yes_no(merged["rule_R5"])
    merged["Bursty Submission Timing"] = _yes_no(merged["rule_R6"])
    merged["Branching / Family Logic Issue"] = _yes_no(merged["rule_R8"])
    merged["Impossible Demographics"] = _yes_no(merged["rule_R9"])
    merged["Extreme Fast (R1 + R2)"] = _yes_no(merged["rule_R1"] & merged["rule_R2"])
    merged["Status-Quo Cluster"] = merged["cluster_status_quo"].fillna("Not clustered")
    merged["Study"] = merged["source_project"].map(PROJECT_LABELS).fillna(merged["source_project"])
    merged["Category"] = merged["classification"].map(
        {
            "Confirmed Bot": "Confirmed Bot",
            "Needs Review": "Needs Human Review",
            "Likely Real Caregiver": "Likely Real Caregivers",
            "Real Caregiver": "Real Caregivers",
        }
    ).fillna(merged["classification"])

    stakeholder_records = merged[WORKBOOK_EXPORT_COLUMNS + ["Category", "source_project", "classification"]].copy()
    stakeholder_records["Record ID"] = stakeholder_records["Record ID"].astype(str)
    return stakeholder_records


def _workflow_counts(stakeholder_records: pd.DataFrame) -> dict[str, int]:
    dirty = stakeholder_records.loc[stakeholder_records["source_project"].eq("dirty_4581")].copy()
    hard_failed = dirty["Trust Tier"].eq("Confirmed invalid")
    confirmed_bots = dirty["classification"].eq("Confirmed Bot")
    review_records = dirty["classification"].eq("Needs Review")
    extreme_fast = dirty["Extreme Fast (R1 + R2)"].eq("Yes")
    other_hard = dirty[[
        "Duplicate Response Pattern",
        "Branching / Family Logic Issue",
        "Impossible Demographics",
    ]].eq("Yes").any(axis=1)

    return {
        "study2_total": int(len(dirty)),
        "hard_failed": int(hard_failed.sum()),
        "hard_review": int((review_records & hard_failed).sum()),
        "confirmed_bots": int(confirmed_bots.sum()),
        "soft_review": int((review_records & ~hard_failed).sum()),
        "low_risk": int(dirty["classification"].eq("Likely Real Caregiver").sum()),
        "cleared": int(dirty["classification"].eq("Real Caregiver").sum()),
        "extreme_fast": int(extreme_fast.sum()),
        "extreme_fast_only_speed": int((extreme_fast & ~other_hard).sum()),
    }


def build_workflow_summary(stakeholder_records: pd.DataFrame) -> pd.DataFrame:
    """Summarise the current Study 2 screening workflow in plain English."""
    counts = _workflow_counts(stakeholder_records)
    rows = [
        {
            "Step": "Study 2 raw responses",
            "Study 2 Count": counts["study2_total"],
            "How to read it": "Starting pool for the legacy 4581 review.",
        },
        {
            "Step": "Hard-check failures (R1, R2, R5, R8, R9)",
            "Study 2 Count": counts["hard_failed"],
            "How to read it": "Records that triggered at least one hard check.",
        },
        {
            "Step": "Needs human review from hard checks",
            "Study 2 Count": counts["hard_review"],
            "How to read it": "Hard-check records that are still reviewable rather than fully rejected.",
        },
        {
            "Step": "Confirmed bots / reject",
            "Study 2 Count": counts["confirmed_bots"],
            "How to read it": "Strongest evidence concentration; do not pay unless adjudication reverses it.",
        },
        {
            "Step": "Soft-check review only (no hard check)",
            "Study 2 Count": counts["soft_review"],
            "How to read it": "Two or more soft checks, but no hard-check trigger.",
        },
        {
            "Step": "Low-risk provisional approval",
            "Study 2 Count": counts["low_risk"],
            "How to read it": "Exactly one soft check and no hard checks.",
        },
        {
            "Step": "Cleared for payment now",
            "Study 2 Count": counts["cleared"],
            "How to read it": "Tier 4 pass with no hard or soft screen trigger.",
        },
        {
            "Step": "Extreme fast: broke both R1 and R2",
            "Study 2 Count": counts["extreme_fast"],
            "How to read it": "Cross-cut subset; these records are inside the hard-check path.",
        },
        {
            "Step": "Extreme fast with no other hard check",
            "Study 2 Count": counts["extreme_fast_only_speed"],
            "How to read it": "Fast on both timing rules, but no duplicate or logic hard-check signal.",
        },
    ]
    return pd.DataFrame(rows)


def build_overview_summary(stakeholder_records: pd.DataFrame) -> pd.DataFrame:
    """Create a concise stakeholder count table across both studies."""
    rows: list[dict[str, object]] = []
    for label, mask, interpretation in [
        (
            "Cleared for payment now",
            stakeholder_records["Payment Decision"].eq("Cleared for payment now"),
            "Tier 4 pass. No screening rule fired.",
        ),
        (
            "Low-risk provisional approval",
            stakeholder_records["Payment Decision"].eq("Low-risk provisional approval"),
            "One soft check only. Provisionally safe but still separate from fully clear records.",
        ),
        (
            "Needs human review",
            stakeholder_records["Payment Decision"].eq("Needs human review"),
            "Manual adjudication queue.",
        ),
        (
            "Confirmed bot / reject",
            stakeholder_records["Payment Decision"].eq("Confirmed bot / reject"),
            "Strongest evidence concentration. Hold payment.",
        ),
        (
            "Extreme fast (R1 + R2)",
            stakeholder_records["Extreme Fast (R1 + R2)"].eq("Yes"),
            "Cross-cut subset of the hard-check path; not a separate payment bucket.",
        ),
    ]:
        subset = stakeholder_records.loc[mask]
        rows.append(
            {
                "Stakeholder Group": label,
                "Study 1 (4797)": int(subset["source_project"].eq("clean_4797").sum()),
                "Study 2 (4581)": int(subset["source_project"].eq("dirty_4581").sum()),
                "All Records": int(len(subset)),
                "Interpretation": interpretation,
            }
        )

    r1r2_only = stakeholder_records.loc[
        stakeholder_records["Extreme Fast (R1 + R2)"].eq("Yes")
        & stakeholder_records[[
            "Duplicate Response Pattern",
            "Branching / Family Logic Issue",
            "Impossible Demographics",
        ]].ne("Yes").all(axis=1)
    ]
    rows.append(
        {
            "Stakeholder Group": "Extreme fast with no other hard check",
            "Study 1 (4797)": int(r1r2_only["source_project"].eq("clean_4797").sum()),
            "Study 2 (4581)": int(r1r2_only["source_project"].eq("dirty_4581").sum()),
            "All Records": int(len(r1r2_only)),
            "Interpretation": "Broke both timing rules, but no duplicate or logic hard-check signal.",
        }
    )

    return pd.DataFrame(rows)


def format_workflow_pipeline(stakeholder_records: pd.DataFrame) -> str:
    """Render a compact ASCII workflow diagram with live counts."""
    counts = _workflow_counts(stakeholder_records)
    return "\n".join(
        [
            "┌──────────────────────────────────────────────┐",
            f"│ {counts['study2_total']:>4} Study 2 responses (Project 4581)     │",
            "└──────────────────────┬───────────────────────┘",
            "                       │",
            "                       ▼",
            "┌──────────────────────────────────────────────┐",
            f"│ Hard checks: R1, R2, R5, R8, R9  -> {counts['hard_failed']:>4} │",
            "└───────────────┬──────────────────┬───────────┘",
            "                │                  │",
            "                ▼                  ▼",
            f"      Needs review from hard   Confirmed bot / reject",
            f"      checks: {counts['hard_review']:>4}            {counts['confirmed_bots']:>4}",
            "",
            "┌──────────────────────────────────────────────┐",
            "│ No hard check -> apply soft checks (R3, R4,  │",
            "│ R6, R7)                                      │",
            "└───────────────┬──────────────────┬───────────┘",
            "                │                  │",
            "                ▼                  ▼",
            f"      Needs human review      Low-risk provisional",
            f"      (>=2 soft flags): {counts['soft_review']:>4}     approval: {counts['low_risk']:>4}",
            "",
            f"Cleared for payment now (Tier 4 pass): {counts['cleared']}",
            f"Cross-cut timing note: {counts['extreme_fast']} broke both R1 and R2; {counts['extreme_fast_only_speed']} had no other hard check.",
        ]
    )


def build_demographic_comparison(stakeholder_records: pd.DataFrame) -> pd.DataFrame:
    """Compare the 223 fully cleared records with the 1,048 review records."""
    compare = stakeholder_records.loc[
        stakeholder_records["Payment Decision"].isin(
            ["Cleared for payment now", "Needs human review"]
        )
    ].copy()
    group_map = {
        "Cleared for payment now": "Cleared for payment (223)",
        "Needs human review": "Needs human review (1048)",
    }
    compare["Comparison Group"] = compare["Payment Decision"].map(group_map)

    def share_row(metric: str, mask: pd.Series) -> dict[str, object]:
        valid = compare.loc[mask.index]
        rows: dict[str, object] = {"Metric": metric}
        pct_values: dict[str, float] = {}
        for group in group_map.values():
            subset = valid.loc[valid["Comparison Group"].eq(group)]
            denom = int(len(subset))
            num = int(mask.loc[subset.index].sum())
            pct = (num / denom * 100) if denom else np.nan
            pct_values[group] = pct
            rows[group] = f"{num}/{denom} ({pct:.1f}%)" if denom else "0/0"
        rows["Difference"] = (
            f"{pct_values[group_map['Cleared for payment now']] - pct_values[group_map['Needs human review']]:.1f} pp"
        )
        return rows

    cleared_label = group_map["Cleared for payment now"]
    review_label = group_map["Needs human review"]
    cleared_age = compare.loc[compare["Comparison Group"].eq(cleared_label), "Caregiver Age"].dropna()
    review_age = compare.loc[compare["Comparison Group"].eq(review_label), "Caregiver Age"].dropna()
    cleared_total = int(compare["Comparison Group"].eq(cleared_label).sum())
    review_total = int(compare["Comparison Group"].eq(review_label).sum())
    rows: list[dict[str, object]] = [
        {
            "Metric": "Caregiver age available",
            cleared_label: f"{len(cleared_age)}/{cleared_total} records",
            review_label: f"{len(review_age)}/{review_total} records",
            "Difference": "Study 2 age comparison is not interpretable from the current cache.",
        }
    ]

    rows.extend(
        [
            share_row("Women", compare["Gender"].eq("Woman")),
            share_row("Hispanic/Latino", compare["Ethnicity"].str.contains("Hispanic/Latino", na=False)),
            share_row("White", compare["Race"].str.contains("White", na=False)),
            share_row(
                "Black or African American",
                compare["Race"].str.contains("Black or African American", na=False),
            ),
            share_row(
                "American Indian/Alaska Native",
                compare["Race"].str.contains("American Indian/Alaska Native", na=False),
            ),
        ]
    )

    return pd.DataFrame(rows)


def build_stakeholder_views(
    output_dir: Path,
    cache_dir: Optional[Path] = None,
) -> dict[str, object]:
    """Load the stakeholder-ready tables, records, and derived summaries."""
    stakeholder_records = build_stakeholder_record_view(output_dir, cache_dir)
    master_summary = load_master_summary(output_dir)
    detailed_breakdown = load_detailed_breakdown(output_dir)
    confirmed_bots = load_confirmed_bots(output_dir)
    record_flags = load_record_flags(output_dir)

    return {
        "master_summary": master_summary,
        "detailed_breakdown": detailed_breakdown,
        "confirmed_bots": confirmed_bots,
        "record_flags": record_flags,
        "stakeholder_records": stakeholder_records,
        "cleared_records": stakeholder_records.loc[
            stakeholder_records["Payment Decision"].eq("Cleared for payment now")
        ].copy(),
        "low_risk_records": stakeholder_records.loc[
            stakeholder_records["Payment Decision"].eq("Low-risk provisional approval")
        ].copy(),
        "review_records": stakeholder_records.loc[
            stakeholder_records["Payment Decision"].eq("Needs human review")
        ].copy(),
        "confirmed_bot_records": stakeholder_records.loc[
            stakeholder_records["Payment Decision"].eq("Confirmed bot / reject")
        ].copy(),
        "extreme_fast_records": stakeholder_records.loc[
            stakeholder_records["Extreme Fast (R1 + R2)"].eq("Yes")
        ].copy(),
        "overview_summary": build_overview_summary(stakeholder_records),
        "workflow_summary": build_workflow_summary(stakeholder_records),
        "workflow_diagram": format_workflow_pipeline(stakeholder_records),
        "demographic_comparison": build_demographic_comparison(stakeholder_records),
    }


# ── Plotting functions ──────────────────────────────────────────────────────

def _apply_stakeholder_style(ax: plt.Axes) -> None:
    """Apply clean, large-label formatting for stakeholder presentation."""
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(labelsize=11)
    ax.xaxis.label.set_size(12)
    ax.yaxis.label.set_size(12)
    ax.title.set_size(14)
    ax.title.set_weight("bold")


def plot_category_summary(master_summary: pd.DataFrame, ax: Optional[plt.Axes] = None) -> plt.Figure:
    """Simple bar chart showing record count per category."""
    # Exclude the Total row
    df = master_summary[master_summary["classification"] != "Total"].copy()

    # Map to ordered categories
    cat_map = {
        "Confirmed Bots": "Confirmed Bot",
        "Needs Human Review": "Needs Human Review",
        "Likely Real Caregivers": "Likely Real Caregivers",
        "Real Caregivers": "Real Caregivers",
    }
    df["cat"] = df["classification"].map(cat_map).fillna(df["classification"])

    if ax is None:
        fig, ax = plt.subplots(figsize=(10, 5))
    else:
        fig = ax.get_figure()

    # Plot bars
    colors = [CATEGORY_COLORS.get(c, "#6B7280") for c in df["cat"]]
    bars = ax.barh(df["classification"], df["total_n"], color=colors, edgecolor="white", linewidth=0.5)

    # Add count labels
    for bar, val in zip(bars, df["total_n"]):
        ax.text(bar.get_width() + 8, bar.get_y() + bar.get_height() / 2,
                f"{val:,}", va="center", fontsize=12, fontweight="bold")

    ax.set_xlabel("Number of Records")
    ax.set_title("Record Classification Summary (All Projects)")
    ax.invert_yaxis()
    _apply_stakeholder_style(ax)
    fig.tight_layout()
    return fig


def plot_completion_time_histogram(
    timed_flags: pd.DataFrame,
    time_col: str = "Total Survey Time (min)",
    threshold_min: float = 11.57,
    title: str = "Survey Completion Time Distribution",
    ax: Optional[plt.Axes] = None,
) -> plt.Figure:
    """Histogram of completion times, colour-coded by category."""
    df = timed_flags.dropna(subset=[time_col]).copy()
    # Cap at reasonable range for visibility
    df = df[df[time_col] <= 120]

    if ax is None:
        fig, ax = plt.subplots(figsize=(11, 5))
    else:
        fig = ax.get_figure()

    for cat in CATEGORY_ORDER:
        subset = df[df["Category"] == cat]
        if len(subset) == 0:
            continue
        ax.hist(
            subset[time_col],
            bins=50,
            alpha=0.65,
            label=f"{cat} (n={len(subset)})",
            color=CATEGORY_COLORS.get(cat, "#6B7280"),
            edgecolor="white",
            linewidth=0.3,
        )

    ax.axvline(threshold_min, color="#DC2626", linestyle="--", linewidth=2,
               label=f"Threshold: {threshold_min} min")
    ax.set_xlabel("Time (minutes)")
    ax.set_ylabel("Number of Records")
    ax.set_title(title)
    ax.legend(frameon=False, fontsize=10)
    _apply_stakeholder_style(ax)
    fig.tight_layout()
    return fig


def plot_rule_violation_frequency(
    record_flags: pd.DataFrame,
    project_filter: Optional[str] = None,
    ax: Optional[plt.Axes] = None,
) -> plt.Figure:
    """Horizontal bar chart: which rules fire most often."""
    df = record_flags.copy()
    if project_filter:
        df = df[df["source_project"] == project_filter]

    rule_cols = [c for c in df.columns if c.startswith("rule_R")]
    counts = df[rule_cols].sum().sort_values(ascending=True)
    labels = [RULE_FRIENDLY_NAMES.get(r, r) for r in counts.index]

    if ax is None:
        fig, ax = plt.subplots(figsize=(10, 5.5))
    else:
        fig = ax.get_figure()

    bars = ax.barh(labels, counts.values, color="#1F5A7A", edgecolor="white", linewidth=0.5)
    for bar, val in zip(bars, counts.values):
        ax.text(bar.get_width() + 5, bar.get_y() + bar.get_height() / 2,
                str(int(val)), va="center", fontsize=11)

    project_label = f" ({project_filter})" if project_filter else ""
    ax.set_xlabel("Number of Records Flagged")
    ax.set_title(f"Rule Violation Frequency{project_label}")
    _apply_stakeholder_style(ax)
    fig.tight_layout()
    return fig


def plot_tier_distribution(
    record_flags: pd.DataFrame,
    ax: Optional[plt.Axes] = None,
) -> plt.Figure:
    """Stacked horizontal bar: tier breakdown per project."""
    tier_order = ["Confirmed invalid", "High suspicion", "Uncertain", "Pass"]
    tier_colors = {
        "Confirmed invalid": "#A85D75",
        "High suspicion": "#C67C2D",
        "Uncertain": "#B08A2E",
        "Pass": "#1F5A7A",
    }

    pivot = (
        record_flags.groupby(["source_project", "tier_label"])
        .size()
        .reset_index(name="n")
        .pivot(index="source_project", columns="tier_label", values="n")
        .fillna(0)
    )
    # Normalise to percentages
    pivot_pct = pivot.div(pivot.sum(axis=1), axis=0) * 100
    pivot_pct = pivot_pct.reindex(columns=tier_order, fill_value=0)

    if ax is None:
        fig, ax = plt.subplots(figsize=(10, 4))
    else:
        fig = ax.get_figure()

    left = pd.Series(0.0, index=pivot_pct.index)
    for label in tier_order:
        ax.barh(pivot_pct.index, pivot_pct[label], left=left,
                color=tier_colors[label], label=label, edgecolor="white", linewidth=0.5)
        left += pivot_pct[label]

    ax.set_xlabel("% of Records")
    ax.set_title("Trust Tier Distribution by Project")
    ax.legend(frameon=False, loc="lower right", fontsize=10)
    _apply_stakeholder_style(ax)
    fig.tight_layout()
    return fig


def plot_detailed_breakdown(
    detailed: pd.DataFrame,
    ax: Optional[plt.Axes] = None,
) -> plt.Figure:
    """Horizontal bar chart of the detailed sub-category breakdown."""
    df = detailed[detailed["classification"] != "Total"].copy()

    if ax is None:
        fig, ax = plt.subplots(figsize=(12, 6))
    else:
        fig = ax.get_figure()

    # Build colour list
    cat_colour_map = {
        "Real Caregiver": "#1F5A7A",
        "Likely Real": "#B08A2E",
        "Needs Review": "#C67C2D",
        "Confirmed Bot": "#A85D75",
    }
    colors = [cat_colour_map.get(c, "#6B7280") for c in df["classification"]]

    bars = ax.barh(df["sub_category_reason"], df["total_n"], color=colors,
                   edgecolor="white", linewidth=0.5)
    for bar, val in zip(bars, df["total_n"]):
        ax.text(bar.get_width() + 5, bar.get_y() + bar.get_height() / 2,
                str(int(val)), va="center", fontsize=10)

    ax.set_xlabel("Number of Records")
    ax.set_title("Detailed Record Breakdown by Sub-Category")
    ax.invert_yaxis()
    _apply_stakeholder_style(ax)
    fig.tight_layout()
    return fig


def plot_rule_overlap_heatmap(
    record_flags: pd.DataFrame,
    project_filter: str = "dirty_4581",
    ax: Optional[plt.Axes] = None,
) -> plt.Figure:
    """Simple co-violation heatmap for the hard-check rules."""
    df = record_flags.loc[record_flags["source_project"].eq(project_filter)].copy()
    rule_cols = ["rule_R1", "rule_R2", "rule_R5", "rule_R8", "rule_R9"]
    matrix = df[rule_cols].astype(int).T @ df[rule_cols].astype(int)
    matrix.index = [RULE_FRIENDLY_NAMES[col] for col in matrix.index]
    matrix.columns = [RULE_FRIENDLY_NAMES[col] for col in matrix.columns]

    if ax is None:
        fig, ax = plt.subplots(figsize=(9.8, 7.0))
    else:
        fig = ax.get_figure()

    sns.heatmap(
        matrix,
        annot=True,
        fmt="d",
        cmap="Blues",
        linewidths=0.5,
        square=True,
        cbar_kws={"label": "Co-flagged records"},
        ax=ax,
    )
    ax.set_title("Hard-check overlap in Study 2")
    _apply_stakeholder_style(ax)
    fig.tight_layout()
    return fig


def plot_clear_vs_review_demographics(
    stakeholder_records: pd.DataFrame,
    axes: Optional[tuple[plt.Axes, plt.Axes]] = None,
) -> plt.Figure:
    """Compare cleared-for-payment records with the human-review queue."""
    compare = stakeholder_records.loc[
        stakeholder_records["Payment Decision"].isin(
            ["Cleared for payment now", "Needs human review"]
        )
    ].copy()
    compare["Comparison Group"] = compare["Payment Decision"].map(
        {
            "Cleared for payment now": "Cleared for payment",
            "Needs human review": "Needs human review",
        }
    )
    cleared_group = "Cleared for payment"
    review_group = "Needs human review"

    if axes is None:
        fig, axes = plt.subplots(1, 2, figsize=(15, 5.5))
    else:
        fig = axes[0].get_figure()

    age_ax, share_ax = axes
    age_counts = pd.DataFrame(
        [
            {
                "Comparison Group": cleared_group,
                "Records with usable age": int(
                    compare.loc[compare["Comparison Group"].eq(cleared_group), "Caregiver Age"].notna().sum()
                ),
                "Total records": int(compare["Comparison Group"].eq(cleared_group).sum()),
            },
            {
                "Comparison Group": review_group,
                "Records with usable age": int(
                    compare.loc[compare["Comparison Group"].eq(review_group), "Caregiver Age"].notna().sum()
                ),
                "Total records": int(compare["Comparison Group"].eq(review_group).sum()),
            },
        ]
    )
    sns.barplot(
        data=age_counts,
        x="Comparison Group",
        y="Records with usable age",
        hue="Comparison Group",
        order=[cleared_group, review_group],
        palette={
            cleared_group: CATEGORY_COLORS["Real Caregivers"],
            review_group: CATEGORY_COLORS["Needs Human Review"],
        },
        dodge=False,
        ax=age_ax,
    )
    if age_ax.legend_ is not None:
        age_ax.legend_.remove()
    for index, row in age_counts.iterrows():
        age_ax.text(
            index,
            row["Records with usable age"] + max(age_counts["Records with usable age"].max() * 0.03, 1),
            f"{int(row['Records with usable age'])}/{int(row['Total records'])}",
            ha="center",
            fontsize=10,
        )
    age_ax.text(
        0.5,
        0.96,
        "Study 2 currently lacks usable caregiver-age data,\nso this panel shows age coverage instead of age distribution.",
        ha="center",
        va="top",
        transform=age_ax.transAxes,
        fontsize=10,
    )
    age_ax.set_title("Caregiver age coverage")
    age_ax.set_xlabel("")
    age_ax.set_ylabel("Records with usable age")
    _apply_stakeholder_style(age_ax)

    share_rows: list[dict[str, object]] = []
    for metric, mask in [
        ("Women", compare["Gender"].eq("Woman")),
        ("Hispanic/Latino", compare["Ethnicity"].str.contains("Hispanic/Latino", na=False)),
        ("White", compare["Race"].str.contains("White", na=False)),
        (
            "Black or African American",
            compare["Race"].str.contains("Black or African American", na=False),
        ),
        (
            "American Indian/Alaska Native",
            compare["Race"].str.contains("American Indian/Alaska Native", na=False),
        ),
    ]:
        for group in [cleared_group, review_group]:
            subset = compare.loc[compare["Comparison Group"].eq(group)]
            pct = mask.loc[subset.index].mean() * 100 if len(subset) else np.nan
            share_rows.append({"Metric": metric, "Group": group, "Percent": pct})

    share_df = pd.DataFrame(share_rows)
    sns.barplot(
        data=share_df,
        x="Percent",
        y="Metric",
        hue="Group",
        palette={
            cleared_group: CATEGORY_COLORS["Real Caregivers"],
            review_group: CATEGORY_COLORS["Needs Human Review"],
        },
        ax=share_ax,
    )
    share_ax.set_title("Demographic shares: cleared vs review")
    share_ax.set_xlabel("Percent of records in each group")
    share_ax.set_ylabel("")
    share_ax.legend(frameon=False)
    _apply_stakeholder_style(share_ax)
    fig.tight_layout()
    return fig


# ── Pipeline text diagram ───────────────────────────────────────────────────

WORKFLOW_PIPELINE = """
┌──────────────────────────────────┐
│  1,779 Raw Responses             │
│  (Study 2 — Project 4581)        │
└───────────────┬──────────────────┘
                │
                ▼
┌──────────────────────────────────┐
│  STEP 1: Apply Hard Checks       │
│  (Tier 1 Rules)                   │
│                                   │
│  R1: Full survey < 11.57 min      │
│  R2: Attitudes section < 7.85 min │
│  R5: Duplicate response pattern   │
│  R8: Illogical family info        │
│  R9: Impossible demographics      │
└───────────────┬──────────────────┘
                │
        ┌───────┴───────┐
        │ Any Hard      │
        │ Check Failed? │
        └──┬─────────┬──┘
        YES│         │NO
           ▼         ▼
   ┌───────────┐  ┌──────────────────────────────┐
   │ FLAGGED   │  │  STEP 2: Apply Soft Checks    │
   │ (Tier 1)  │  │  (Tier 2 / 3 / 4 Rules)       │
   │           │  │                                │
   │ 508       │  │  R3: Section below speed floor  │
   │ records   │  │  R4: Flat-line responses        │
   └─────┬─────┘  │  R6: Bursty submission timing   │
         │        │  R7: Near-duplicate open text    │
         │        └──────────────┬─────────────────┘
         │               ┌──────┴──────┐
         │               │ ≥2 Soft     │
         │               │ Flags?      │
         │               └──┬──────┬───┘
         │               YES│      │NO
         │                  ▼      ▼
         │         ┌──────────┐  ┌─────────────────────────┐
         │         │ Tier 2:  │  │ Tier 3/4:               │
         │         │ Needs    │  │ Likely Real / Real       │
         │         │ Review   │  │ Caregivers               │
         │         │ (563)    │  │ (708 = 679 likely + 70   │
         │         └──────────┘  │  real from dirty)        │
         │                       └─────────────────────────┘
         ▼
   ┌──────────────────────┐
   │ ≥3 hard hits OR      │
   │ R8/R9 + speed flag?  │
   └──┬────────────┬──────┘
   YES│            │NO
      ▼            ▼
┌───────────┐  ┌─────────────────┐
│ Confirmed │  │ Single / double │
│ Bot (6)   │  │ hard flag →     │
│           │  │ Needs Review    │
│ REJECT    │  │ (502)           │
└───────────┘  └─────────────────┘
"""


def print_workflow_pipeline() -> None:
    """Print the screening workflow as a text diagram."""
    print(WORKFLOW_PIPELINE)


# ── Excel export ────────────────────────────────────────────────────────────

def _friendly_rename(df: pd.DataFrame) -> pd.DataFrame:
    """Rename columns using friendly names, keeping unmapped columns as-is."""
    return df.rename(columns=COLUMN_FRIENDLY_NAMES)


def _autosize_sheet(worksheet, wrap_text: bool = False) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    if wrap_text:
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = cell.alignment.copy(wrap_text=True)


def export_stakeholder_excel(
    output_dir: Path,
    cache_dir: Optional[Path] = None,
    excel_filename: str = "ESD_Bot_Analysis_Stakeholder_Summary.xlsx",
) -> Path:
    """Write a multi-sheet Excel workbook for stakeholder review.

    Sheets:
        1. Overview — main counts used in meetings
        2. Workflow Summary — step-by-step Study 2 screening flow
        3. All Records — stakeholder-friendly record-level export
        4. Cleared Now — 223 fully cleared records
        5. Low-Risk Approval — one-soft-flag records
        6. Needs Review — manual adjudication queue
        7. Extreme Fast R1R2 — records that broke both timing rules
        8. Confirmed Bots — the 6 strongest reject cases
        9. 223 vs 1048 Demo — simple cleared-vs-review comparison
        10. Rule Guide — plain-English rule definitions
        11. Pipeline Diagram — text workflow diagram
    """
    views = build_stakeholder_views(output_dir, cache_dir)
    rules = load_rule_definitions(output_dir)

    excel_path = output_dir / excel_filename

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        views["overview_summary"].to_excel(writer, sheet_name="Overview", index=False)
        views["workflow_summary"].to_excel(writer, sheet_name="Workflow Summary", index=False)
        views["stakeholder_records"][WORKBOOK_EXPORT_COLUMNS].to_excel(
            writer,
            sheet_name="All Records",
            index=False,
        )
        views["cleared_records"][WORKBOOK_EXPORT_COLUMNS].to_excel(
            writer,
            sheet_name="Cleared Now",
            index=False,
        )
        views["low_risk_records"][WORKBOOK_EXPORT_COLUMNS].to_excel(
            writer,
            sheet_name="Low-Risk Approval",
            index=False,
        )
        views["review_records"][WORKBOOK_EXPORT_COLUMNS].to_excel(
            writer,
            sheet_name="Needs Review",
            index=False,
        )
        views["extreme_fast_records"][WORKBOOK_EXPORT_COLUMNS].to_excel(
            writer,
            sheet_name="Extreme Fast R1R2",
            index=False,
        )
        views["confirmed_bot_records"][WORKBOOK_EXPORT_COLUMNS].to_excel(
            writer,
            sheet_name="Confirmed Bots",
            index=False,
        )
        views["demographic_comparison"].to_excel(writer, sheet_name="223 vs 1048 Demo", index=False)

        rules_sheet = rules[["rule", "Friendly Name", "definition", "threshold", "justification"]].copy()
        rules_sheet.columns = ["Rule ID", "Rule Name", "Description", "Threshold", "Justification"]
        rules_sheet.to_excel(writer, sheet_name="Rule Guide", index=False)

        pipeline_df = pd.DataFrame({
            "Screening Pipeline": views["workflow_diagram"].split("\n")
        })
        pipeline_df.to_excel(writer, sheet_name="Pipeline Diagram", index=False)

        _autosize_sheet(writer.book["Overview"], wrap_text=True)
        _autosize_sheet(writer.book["Workflow Summary"], wrap_text=True)
        _autosize_sheet(writer.book["All Records"])
        _autosize_sheet(writer.book["Cleared Now"])
        _autosize_sheet(writer.book["Low-Risk Approval"])
        _autosize_sheet(writer.book["Needs Review"])
        _autosize_sheet(writer.book["Extreme Fast R1R2"])
        _autosize_sheet(writer.book["Confirmed Bots"])
        _autosize_sheet(writer.book["223 vs 1048 Demo"], wrap_text=True)
        _autosize_sheet(writer.book["Rule Guide"], wrap_text=True)
        writer.book["Pipeline Diagram"].column_dimensions["A"].width = 72

    return excel_path
