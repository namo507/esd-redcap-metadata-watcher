#!/usr/bin/env python
"""
NANO Family Income Report — Simplified 3-Sheet Version
=======================================================
Sheet 1: Consent Data       — All participants' consent-only income & employment + histogram
Sheet 2: Follow-Up Data     — 12/24/36-month entries (separate from consent, not merged)
Sheet 3: Analysis Summary   — Stakeholder-friendly stability & trend metrics
"""

import os
from collections import OrderedDict

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─── Configuration ───────────────────────────────────────────────────────────
API_URL = "https://redcap.research.sc.edu/api/"
TOKEN = "6324B3FAA4E18D8D513776801CFABA20"
FORM_NAME = "family_information_form"

DATE_FIELD = "fif_doe"
INCOME_FIELD = "fif_income"
PRIMARY_EMPLOYMENT_FIELD = "fif_cg1employment"
SECONDARY_EMPLOYMENT_FIELD = "fif_cg2employment"

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_FILENAME = os.path.join(OUTPUT_DIR, "NANO_Family_Income_Report.xlsx")

INCOME_ORDER = [
    "$20,000 or less",
    "$20,001 - $40,000",
    "$40,001 - $60,000",
    "$60,001 - $80,000",
    "$80,001 - $100,000",
    "$100,001 - $125,000",
    "$125,001 - $150,000",
    "$150,001 - $200,000",
    "$200,001 or higher",
]
# Short labels for histogram x-axis
INCOME_SHORT = [
    "<$20K", "$20-40K", "$40-60K", "$60-80K", "$80-100K",
    "$100-125K", "$125-150K", "$150-200K", ">$200K",
]
NON_INCOME_LABELS = ["I don't know", "I prefer not to answer"]

# Column names exactly as requested
COL_ID = "Participant ID Number"
COL_INCOME = "Total Income"
COL_PRIMARY = "Primary Caregiver - Employment Status"
COL_SECONDARY = "Secondary Caregiver - Employment Status"
DATA_COLUMNS = [COL_ID, COL_INCOME, COL_PRIMARY, COL_SECONDARY]

# ─── Styling ─────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEAA7", end_color="FFEAA7", fill_type="solid")
LIGHT_GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(italic=True, size=10, color="666666")
SECTION_FONT = Font(bold=True, size=12, color="1F4E78")


# ─── Helpers ─────────────────────────────────────────────────────────────────
def redcap_post(payload: dict) -> list:
    resp = requests.post(
        API_URL, data={"token": TOKEN, "returnFormat": "json", **payload}, timeout=120
    )
    resp.raise_for_status()
    data = resp.json()
    if isinstance(data, dict) and data.get("error"):
        raise RuntimeError(f"REDCap API error: {data['error']}")
    return data


def clean(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def checkbox_or_single(frame, field_name):
    cb_cols = [c for c in frame.columns if c.startswith(f"{field_name}___")]
    if field_name in frame.columns:
        return frame[field_name].map(clean)
    if cb_cols:
        def _labels(row):
            sel = []
            for col in cb_cols:
                v = clean(row.get(col, ""))
                if v and v.lower() not in {"0", "unchecked", "false", "no"}:
                    sel.append(v if v.lower() not in {"1", "checked", "true", "yes"} else col.split("___", 1)[1])
            return "; ".join(sel)
        return frame.apply(_labels, axis=1)
    return pd.Series([""] * len(frame), index=frame.index)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def auto_width(ws, ncols, max_w=50, min_w=14):
    for j in range(1, ncols + 1):
        lens = [len(str(ws.cell(row=r, column=j).value or "")) for r in range(1, ws.max_row + 1)]
        ws.column_dimensions[get_column_letter(j)].width = min(max(max(lens) + 3 if lens else 10, min_w), max_w)


def write_data_rows(ws, df, start_row, columns):
    """Write a dataframe's rows into the worksheet starting at start_row."""
    for i, (_, r) in enumerate(df.iterrows()):
        rn = start_row + i
        for j, col in enumerate(columns, 1):
            val = r[col] if r[col] != "" else None
            cell = ws.cell(row=rn, column=j, value=val)
            cell.border = BORDER
            cell.alignment = CENTER if j == 1 else LEFT_WRAP


# ─── Main ────────────────────────────────────────────────────────────────────
def main():
    print("Connecting to REDCap…")
    pj = requests.post(API_URL, data={
        "token": TOKEN, "content": "project", "format": "json", "returnFormat": "json"
    }, timeout=30)
    pj.raise_for_status()
    print(f"HTTP Status: {pj.status_code}")
    print(f"Project: {pj.json().get('project_title', '?')}")

    meta = pd.DataFrame(redcap_post({"content": "metadata", "format": "json"}))
    REC_ID = meta.iloc[0]["field_name"]

    raw = pd.DataFrame(redcap_post({
        "content": "record", "format": "json", "type": "flat",
        "rawOrLabel": "label", "rawOrLabelHeaders": "raw",
        "fields[0]": REC_ID, "fields[1]": DATE_FIELD,
        "fields[2]": INCOME_FIELD, "fields[3]": PRIMARY_EMPLOYMENT_FIELD,
        "fields[4]": SECONDARY_EMPLOYMENT_FIELD,
        "forms[0]": FORM_NAME, "exportCheckboxLabel": "true",
    }))
    print(f"Raw rows: {len(raw)}")

    # Build long table
    long = pd.DataFrame({
        COL_ID: raw[REC_ID].map(clean),
        "Date": raw[DATE_FIELD].map(clean) if DATE_FIELD in raw.columns else "",
        COL_INCOME: checkbox_or_single(raw, INCOME_FIELD),
        COL_PRIMARY: checkbox_or_single(raw, PRIMARY_EMPLOYMENT_FIELD),
        COL_SECONDARY: checkbox_or_single(raw, SECONDARY_EMPLOYMENT_FIELD),
    })
    long = long[long[COL_ID].ne("")]
    long = long[~long[COL_ID].str.upper().str.contains("TEST")]
    dcols = [COL_INCOME, COL_PRIMARY, COL_SECONDARY]
    long = long[~long[dcols].apply(lambda r: all(v == "" for v in r), axis=1)]
    long = long.sort_values([COL_ID, "Date"]).reset_index(drop=True)
    long["Event #"] = long.groupby(COL_ID).cumcount() + 1

    # Consent = Event 1, Follow-ups = Event 2+
    consent_df = long[long["Event #"] == 1][DATA_COLUMNS].copy().reset_index(drop=True)
    followup_df = long[long["Event #"] > 1].copy().reset_index(drop=True)

    # Map event numbers to labels
    evt_map = {2: "12-Month", 3: "24-Month", 4: "36-Month"}
    followup_df["Time Point"] = followup_df["Event #"].map(lambda x: evt_map.get(x, f"Event {x}"))

    pids = long[COL_ID].unique().tolist()
    print(f"Participants: {len(pids)}, Consent rows: {len(consent_df)}, Follow-up rows: {len(followup_df)}")

    # =================================================================
    wb = Workbook()

    # ── Sheet 1: Consent Data + Histogram ────────────────────────────
    ws1 = wb.active
    ws1.title = "Consent Data"

    # Headers at row 1
    for j, h in enumerate(DATA_COLUMNS, 1):
        ws1.cell(row=1, column=j, value=h)
    style_header(ws1, 1, 4)

    write_data_rows(ws1, consent_df, 2, DATA_COLUMNS)
    auto_width(ws1, 4)
    ws1.freeze_panes = "A2"

    # Histogram data table (placed to the right of main data)
    hist_col_start = 6  # column F
    ws1.cell(row=1, column=hist_col_start, value="Income Range").font = HEADER_FONT
    ws1.cell(row=1, column=hist_col_start).fill = HEADER_FILL
    ws1.cell(row=1, column=hist_col_start).border = BORDER
    ws1.cell(row=1, column=hist_col_start).alignment = CENTER

    ws1.cell(row=1, column=hist_col_start + 1, value="Count").font = HEADER_FONT
    ws1.cell(row=1, column=hist_col_start + 1).fill = HEADER_FILL
    ws1.cell(row=1, column=hist_col_start + 1).border = BORDER
    ws1.cell(row=1, column=hist_col_start + 1).alignment = CENTER

    income_series = consent_df[COL_INCOME]
    hist_data = OrderedDict()
    for bracket, short in zip(INCOME_ORDER, INCOME_SHORT):
        hist_data[short] = int((income_series == bracket).sum())
    for lbl in NON_INCOME_LABELS:
        hist_data[lbl] = int((income_series == lbl).sum())

    for idx, (label, count) in enumerate(hist_data.items(), start=2):
        c1 = ws1.cell(row=idx, column=hist_col_start, value=label)
        c1.border = BORDER
        c1.alignment = LEFT_WRAP
        c2 = ws1.cell(row=idx, column=hist_col_start + 1, value=count)
        c2.border = BORDER
        c2.alignment = CENTER

    ws1.column_dimensions[get_column_letter(hist_col_start)].width = 24
    ws1.column_dimensions[get_column_letter(hist_col_start + 1)].width = 10

    # Bar chart
    last_hist_row = 1 + len(hist_data)
    chart = BarChart()
    chart.type = "col"
    chart.title = "Consent Income Distribution"
    chart.y_axis.title = "Number of Participants"
    chart.x_axis.title = "Total Income"
    chart.style = 10
    chart.width = 26
    chart.height = 14

    data_ref = Reference(ws1, min_col=hist_col_start + 1, min_row=1, max_row=last_hist_row)
    cats_ref = Reference(ws1, min_col=hist_col_start, min_row=2, max_row=last_hist_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.legend = None
    chart.series[0].graphicalProperties.solidFill = "4472C4"

    chart_anchor = f"{get_column_letter(hist_col_start + 3)}2"
    ws1.add_chart(chart, chart_anchor)

    # ── Sheet 2: Follow-Up Data ──────────────────────────────────────
    ws2 = wb.create_sheet("Follow-Up Data")
    fu_columns = ["Time Point", COL_ID, COL_INCOME, COL_PRIMARY, COL_SECONDARY]

    for j, h in enumerate(fu_columns, 1):
        ws2.cell(row=1, column=j, value=h)
    style_header(ws2, 1, len(fu_columns))

    if len(followup_df) > 0:
        followup_sorted = followup_df.sort_values(["Event #", COL_ID]).reset_index(drop=True)
        write_data_rows(ws2, followup_sorted, 2, fu_columns)
    else:
        ws2.cell(row=2, column=1, value="No follow-up data available yet.")

    auto_width(ws2, len(fu_columns))
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Analysis Summary ────────────────────────────────────
    ws3 = wb.create_sheet("Analysis")
    row = 1

    def write_title(text):
        nonlocal row
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws3.cell(row=row, column=1, value=text).font = TITLE_FONT
        row += 1

    def write_subtitle(text):
        nonlocal row
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws3.cell(row=row, column=1, value=text).font = SUBTITLE_FONT
        row += 1

    def write_section(text):
        nonlocal row
        row += 1
        ws3.cell(row=row, column=1, value=text).font = SECTION_FONT
        row += 1

    def write_metric(label, value, fill=None):
        nonlocal row
        c1 = ws3.cell(row=row, column=1, value=label)
        c1.font = Font(size=10)
        c1.border = BORDER
        c2 = ws3.cell(row=row, column=2, value=value)
        c2.font = Font(bold=True, size=10)
        c2.border = BORDER
        if fill:
            c2.fill = fill
        row += 1

    def write_note(text, fill=None):
        nonlocal row
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws3.cell(row=row, column=1, value=text)
        c.font = Font(size=10)
        if fill:
            c.fill = fill
        row += 1

    # Title
    write_title("NANO Family Income — Analysis Summary")
    write_subtitle("Income & employment stability across consent and follow-up time points")

    # ── Section: Consent Overview ────
    write_section("A. CONSENT INCOME OVERVIEW")

    consent_valid = consent_df[consent_df[COL_INCOME].isin(INCOME_ORDER)]
    consent_unknown = consent_df[consent_df[COL_INCOME].isin(NON_INCOME_LABELS)]
    consent_blank = consent_df[~consent_df[COL_INCOME].isin(INCOME_ORDER + NON_INCOME_LABELS)]

    write_metric("Total participants at consent", len(consent_df))
    write_metric("With valid income bracket", len(consent_valid))
    write_metric("Responded 'I don't know' / 'Prefer not to answer'", len(consent_unknown))
    write_metric("Blank / missing income", len(consent_blank))

    # Consent income distribution table
    row += 1
    ws3.cell(row=row, column=1, value="Income Bracket").font = HEADER_FONT
    ws3.cell(row=row, column=1).fill = HEADER_FILL
    ws3.cell(row=row, column=1).border = BORDER
    ws3.cell(row=row, column=2, value="Count").font = HEADER_FONT
    ws3.cell(row=row, column=2).fill = HEADER_FILL
    ws3.cell(row=row, column=2).border = BORDER
    ws3.cell(row=row, column=3, value="%").font = HEADER_FONT
    ws3.cell(row=row, column=3).fill = HEADER_FILL
    ws3.cell(row=row, column=3).border = BORDER
    row += 1

    total_consent = len(consent_df)
    for bracket, short in zip(INCOME_ORDER, INCOME_SHORT):
        cnt = int((consent_df[COL_INCOME] == bracket).sum())
        pct = f"{cnt / total_consent * 100:.1f}%" if total_consent else "0%"
        ws3.cell(row=row, column=1, value=f"{short}  ({bracket})").border = BORDER
        ws3.cell(row=row, column=2, value=cnt).border = BORDER
        ws3.cell(row=row, column=2).alignment = CENTER
        ws3.cell(row=row, column=3, value=pct).border = BORDER
        ws3.cell(row=row, column=3).alignment = CENTER
        row += 1

    # ── Section: Consent Employment Breakdown ────
    write_section("B. CONSENT EMPLOYMENT BREAKDOWN")

    for col_name, label in [(COL_PRIMARY, "Primary Caregiver"), (COL_SECONDARY, "Secondary Caregiver")]:
        ws3.cell(row=row, column=1, value=f"{label} Employment").font = Font(bold=True, size=10, color="1F4E78")
        row += 1
        emp_counts = consent_df[col_name].value_counts()
        for status, cnt in emp_counts.items():
            if status == "":
                status = "(blank / not provided)"
            ws3.cell(row=row, column=1, value=status).border = BORDER
            ws3.cell(row=row, column=2, value=cnt).border = BORDER
            ws3.cell(row=row, column=2).alignment = CENTER
            pct = f"{cnt / total_consent * 100:.1f}%"
            ws3.cell(row=row, column=3, value=pct).border = BORDER
            ws3.cell(row=row, column=3).alignment = CENTER
            row += 1
        row += 1

    # ── Section: Longitudinal Trends ────
    write_section("C. LONGITUDINAL INCOME TRENDS")

    # Count participants per time point
    max_ev = int(long["Event #"].max()) if len(long) > 0 else 1
    event_labels_map = {1: "Consent", 2: "12-Month", 3: "24-Month", 4: "36-Month"}

    ws3.cell(row=row, column=1, value="Time Point").font = HEADER_FONT
    ws3.cell(row=row, column=1).fill = HEADER_FILL
    ws3.cell(row=row, column=1).border = BORDER
    ws3.cell(row=row, column=2, value="Participants").font = HEADER_FONT
    ws3.cell(row=row, column=2).fill = HEADER_FILL
    ws3.cell(row=row, column=2).border = BORDER
    ws3.cell(row=row, column=3, value="With Valid Income").font = HEADER_FONT
    ws3.cell(row=row, column=3).fill = HEADER_FILL
    ws3.cell(row=row, column=3).border = BORDER
    row += 1

    for ev in range(1, max_ev + 1):
        ev_data = long[long["Event #"] == ev]
        ev_valid = ev_data[ev_data[COL_INCOME].isin(INCOME_ORDER)]
        label = event_labels_map.get(ev, f"Event {ev}")
        ws3.cell(row=row, column=1, value=label).border = BORDER
        ws3.cell(row=row, column=2, value=len(ev_data)).border = BORDER
        ws3.cell(row=row, column=2).alignment = CENTER
        ws3.cell(row=row, column=3, value=len(ev_valid)).border = BORDER
        ws3.cell(row=row, column=3).alignment = CENTER
        row += 1

    # ── Section: Income Change Analysis ────
    write_section("D. INCOME STABILITY ASSESSMENT")
    write_note("Comparing consent income to the latest available follow-up for each participant:")
    row += 1

    stable = 0
    increased = 0
    decreased = 0
    unknown_change = 0
    consent_only_count = 0
    change_details = []  # for the per-participant table

    for pid in pids:
        sub = long[long[COL_ID] == pid].sort_values("Event #")
        if sub["Event #"].max() <= 1:
            consent_only_count += 1
            continue
        valid_entries = sub[sub[COL_INCOME].isin(INCOME_ORDER)]
        if len(valid_entries) < 2:
            unknown_change += 1
            continue
        first_income = valid_entries.iloc[0][COL_INCOME]
        last_income = valid_entries.iloc[-1][COL_INCOME]
        first_idx = INCOME_ORDER.index(first_income)
        last_idx = INCOME_ORDER.index(last_income)
        last_ev = int(valid_entries.iloc[-1]["Event #"])
        last_label = event_labels_map.get(last_ev, f"Event {last_ev}")

        if first_idx == last_idx:
            stable += 1
            change_details.append((pid, first_income, last_income, last_label, "Stable", 0))
        elif last_idx > first_idx:
            increased += 1
            change_details.append((pid, first_income, last_income, last_label, "Increased", last_idx - first_idx))
        else:
            decreased += 1
            change_details.append((pid, first_income, last_income, last_label, "Decreased", last_idx - first_idx))

    total_assessed = stable + increased + decreased
    write_metric("Participants with consent only (no follow-up)", consent_only_count)
    write_metric("Participants assessed for change", total_assessed)
    write_metric("Could not assess (missing valid income at both points)", unknown_change)
    row += 1

    ws3.cell(row=row, column=1, value="Result").font = HEADER_FONT
    ws3.cell(row=row, column=1).fill = HEADER_FILL
    ws3.cell(row=row, column=1).border = BORDER
    ws3.cell(row=row, column=2, value="Count").font = HEADER_FONT
    ws3.cell(row=row, column=2).fill = HEADER_FILL
    ws3.cell(row=row, column=2).border = BORDER
    ws3.cell(row=row, column=3, value="% of Assessed").font = HEADER_FONT
    ws3.cell(row=row, column=3).fill = HEADER_FILL
    ws3.cell(row=row, column=3).border = BORDER
    row += 1

    for label, cnt, fill in [
        ("Income STABLE", stable, LIGHT_GREEN),
        ("Income INCREASED", increased, GREEN_FILL),
        ("Income DECREASED", decreased, RED_FILL),
    ]:
        pct = f"{cnt / total_assessed * 100:.1f}%" if total_assessed else "0%"
        c1 = ws3.cell(row=row, column=1, value=label)
        c1.border = BORDER
        c1.fill = fill
        c2 = ws3.cell(row=row, column=2, value=cnt)
        c2.border = BORDER
        c2.alignment = CENTER
        c2.fill = fill
        c3 = ws3.cell(row=row, column=3, value=pct)
        c3.border = BORDER
        c3.alignment = CENTER
        c3.fill = fill
        row += 1

    # ── Section: Per-participant change details (only those who changed) ──
    changed_details = [d for d in change_details if d[4] != "Stable"]
    if changed_details:
        write_section("E. PARTICIPANTS WITH INCOME CHANGES")
        write_note("Only participants whose income bracket changed between consent and follow-up:")
        row += 1

        chg_headers = ["Participant ID", "Consent Income", "Latest Income", "Latest Time Point", "Direction", "Bracket Shift"]
        for j, h in enumerate(chg_headers, 1):
            ws3.cell(row=row, column=j, value=h)
        style_header(ws3, row, len(chg_headers))
        row += 1

        for pid, first_inc, last_inc, tp, direction, shift in sorted(changed_details, key=lambda x: x[0]):
            ws3.cell(row=row, column=1, value=pid).border = BORDER
            ws3.cell(row=row, column=1).alignment = CENTER
            ws3.cell(row=row, column=2, value=first_inc).border = BORDER
            ws3.cell(row=row, column=3, value=last_inc).border = BORDER
            ws3.cell(row=row, column=4, value=tp).border = BORDER
            ws3.cell(row=row, column=4).alignment = CENTER
            c_dir = ws3.cell(row=row, column=5, value=direction)
            c_dir.border = BORDER
            c_dir.alignment = CENTER
            c_dir.fill = GREEN_FILL if direction == "Increased" else RED_FILL
            c_shift = ws3.cell(row=row, column=6, value=shift)
            c_shift.border = BORDER
            c_shift.alignment = CENTER
            row += 1

    # ── Section: Final Verdict ────
    write_section("F. KEY FINDINGS & RECOMMENDATIONS")

    stability_pct = (stable / total_assessed * 100) if total_assessed else 0

    if stability_pct >= 70:
        write_note(f"✅ {stability_pct:.0f}% of families with follow-up have STABLE income across time points.", LIGHT_GREEN)
        write_note("✅ Consent-only income CAN reliably represent household demographics for this cohort.", LIGHT_GREEN)
    elif stability_pct >= 50:
        write_note(f"⚠️ {stability_pct:.0f}% stable — moderate reliability. Consent data is a reasonable approximation.", YELLOW_FILL)
        write_note("⚠️ Note this as a limitation in publications using consent-only demographics.", YELLOW_FILL)
    else:
        write_note(f"❌ Only {stability_pct:.0f}% stable — consent-only income may NOT reliably represent demographics.", RED_FILL)
        write_note("❌ Recommend using longitudinal income data for demographic reporting.", RED_FILL)

    if increased > decreased:
        write_note(f"📈 Trend: More families saw income INCREASE ({increased}) than decrease ({decreased}).")
    elif decreased > increased:
        write_note(f"📉 Trend: More families saw income DECREASE ({decreased}) than increase ({increased}).")
    else:
        write_note(f"↔️ Equal increases ({increased}) and decreases ({decreased}).")

    write_note(f"📊 {consent_only_count} of {len(pids)} participants have consent data only — no follow-up to compare.")

    auto_width(ws3, 6, max_w=55)

    # ── Save ─────────────────────────────────────────────────────────
    wb.save(XLSX_FILENAME)
    print(f"\n✅ Saved: {XLSX_FILENAME}")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
