"""Canonical, read-only NANO/NICO recruitment report generator.

Credentials are loaded only from the process environment or an ignored local
``.env`` file. Aggregate reports are written to ``recruitment_outputs``;
participant-level audit rows are written only to ``recruitment_audit_secure``.
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from html import escape
import json
import os
from pathlib import Path
import re
import tempfile
from typing import Any, Callable, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from redcap import Project

from redcap_client import sanitize_error


DEFAULT_API_URL = "https://redcap.research.sc.edu/api/"
TOKEN_PATTERN = re.compile(r"^[A-Fa-f0-9]{32}$")
FORMULA_PREFIXES = ("=", "+", "-", "@")
PUBLIC_SMALL_CELL_THRESHOLD = 10

CATEGORIES = ("Total", "Minority", "Hispanic")
CATEGORY_LABELS = {
    "Total": "Total Recruitment",
    "Minority": "Racial Minority Recruitment",
    "Hispanic": "Hispanic Ethnicity Recruitment",
}
MONTH_LABELS = {4: "Apr 1", 8: "Aug 1", 12: "Dec 1"}

BLUE = "4F81BD"
NAVY = "17365D"
LIGHT_BLUE = "DCE6F1"
ORANGE = "F6A821"
LIGHT_ORANGE = "FCE4D6"
GREY = "D9E1F2"
LIGHT_GREY = "F3F5F7"
GREEN = "E2F0D9"
RED = "F4CCCC"
WHITE = "FFFFFF"


@dataclass(frozen=True)
class StudyConfig:
    key: str
    label: str
    expected_project_id: int
    record_id: str
    grant: str
    study_title: str
    race_field: str
    race_minority_codes: tuple[int, ...]
    race_white_codes: tuple[int, ...]
    race_unknown_codes: tuple[int, ...]
    race_hispanic_codes: tuple[int, ...]
    ethnicity_field: str
    ethnicity_hispanic_codes: tuple[int, ...]
    ethnicity_unknown_codes: tuple[int, ...]
    secondary_ethnicity: tuple[tuple[str, tuple[int, ...]], ...]
    status_fields: tuple[str, ...]
    exclusion_fields: tuple[str, ...]
    review_fields: tuple[str, ...]
    dual_field: str | None
    date_candidates: tuple[tuple[str, str], ...]
    milestone_start: date
    milestone_end: date
    milestone_months: tuple[int, ...]
    historical_actuals: Mapping[str, tuple[int | str, ...]]
    previous_targets: Mapping[str, tuple[int | None, ...]]
    current_targets: Mapping[str, tuple[int | None, ...]]
    targets_state: str
    targets_source: str
    demographics_state: str
    demographics_note: str


STUDIES: tuple[StudyConfig, ...] = (
    StudyConfig(
        key="NANO",
        label="NANO Study",
        expected_project_id=4218,
        record_id="demo_id",
        grant="MH132925",
        study_title=(
            "The Role of Autonomic Regulation of Attention in the Emergence of ASD"
        ),
        race_field="fif_childrace",
        race_minority_codes=(1, 2, 3, 4),
        race_white_codes=(5,),
        race_unknown_codes=(6,),
        race_hispanic_codes=(),
        ethnicity_field="fif_childethnicity",
        ethnicity_hispanic_codes=(1,),
        ethnicity_unknown_codes=(3,),
        secondary_ethnicity=(),
        status_fields=("demo_ineligible", "demo_unenrolled", "demo_exclude"),
        exclusion_fields=("demo_ineligible", "demo_unenrolled"),
        review_fields=("demo_exclude",),
        dual_field=None,
        date_candidates=(
            ("visit_date", "Visit date"),
            ("bsrc_doe", "Data-sharing consent date"),
            ("papf_parent_date", "Optional media/advertising form date"),
            ("fif_doe", "Family information evaluation date"),
        ),
        milestone_start=date(2024, 8, 1),
        milestone_end=date(2028, 12, 1),
        milestone_months=(4, 8, 12),
        historical_actuals={
            "Total": (63, 108, 128, 151, 172, 219),
            # Secondary suppression prevents adjacent cumulative milestones
            # from revealing an interval count below the public threshold.
            "Minority": (25, 48, 59, 84, "90–99", "90–99"),
            "Hispanic": ("<10", "<10", "<10", "<10", "10–19", "10–19"),
        },
        previous_targets={
            "Total": (90, 110, 130, 150, 170, 190, 200, None, None, None, 160),
            "Minority": (36, 44, 52, 60, 68, 76, 84, None, None, None, 32),
            "Hispanic": (3, 5, 7, 9, 11, 13, 14, None, None, None, 7),
        },
        current_targets={
            "Total": (5, 10, 110, 120, 130, 140, 150, 160, 170, 180, 190, 200),
            "Minority": (1, 2, None, None, None, None, None, None, None, None, None, 40),
            "Hispanic": (None, None, 1, None, 1, None, 1, None, 1, None, 1, 10),
        },
        targets_state="provisional / unverified",
        targets_source="Prior reporting configuration; not reconciled to the NIH-approved plan",
        demographics_state="configured; pending study-owner policy confirmation",
        demographics_note=(
            "Codes match the exported data dictionary; minority/Hispanic reporting "
            "rules still require study-owner approval"
        ),
    ),
    StudyConfig(
        key="NICO",
        label="NICO Study (NICU study)",
        expected_project_id=3836,
        record_id="id",
        grant="Pending",
        study_title="NICO Study",
        race_field="race",
        race_minority_codes=(1, 2, 3, 5),
        race_white_codes=(6,),
        race_unknown_codes=(7,),
        race_hispanic_codes=(4,),
        ethnicity_field="fif_childethnicity",
        ethnicity_hispanic_codes=(1,),
        ethnicity_unknown_codes=(3,),
        secondary_ethnicity=(("ethnicity", (0,)),),
        status_fields=("demo_ineligible", "demo_unenrolled", "demo_exclude"),
        exclusion_fields=("demo_ineligible", "demo_unenrolled"),
        review_fields=("demo_exclude",),
        dual_field="dual_enrolled",
        date_candidates=(
            ("visit_date", "Visit date"),
            ("dob", "Infant date of birth (not enrollment)"),
        ),
        milestone_start=date(2024, 8, 1),
        milestone_end=date(2028, 12, 1),
        milestone_months=(4, 8, 12),
        historical_actuals={},
        previous_targets={},
        current_targets={},
        targets_state="missing",
        targets_source="NIH milestone dates and targets were not supplied",
        demographics_state="provisional",
        demographics_note=(
            "Race code 4 and two ethnicity sources are unioned for observed "
            "Hispanic/Latino; race code 7 is treated as unknown/other"
        ),
    ),
)


@dataclass
class StudyResult:
    config: StudyConfig
    project_info: dict[str, Any]
    metadata_fields: set[str]
    participant_audit: pd.DataFrame
    date_coverage: list[dict[str, Any]]
    live_counts: dict[str, int]
    data_quality: dict[str, int]
    qa_checks: list[dict[str, Any]]
    raw_row_count: int
    strict_eligibility: bool

    @property
    def project_id(self) -> int:
        return int(self.project_info.get("project_id"))


@dataclass
class ReportResult:
    report_date: date
    generated_at: datetime
    counting_policy: str
    studies: dict[str, StudyResult]
    paths: dict[str, Path] = field(default_factory=dict)


def load_env_file(path: str | Path = ".env") -> None:
    """Load a simple dotenv file without overriding existing environment values."""

    env_path = Path(path)
    if not env_path.is_file():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        name, value = line.split("=", 1)
        name = name.strip()
        value = value.strip()
        if value[:1] == value[-1:] and value[:1] in {"'", '"'}:
            value = value[1:-1]
        if name and name not in os.environ:
            os.environ[name] = value


def _required_token(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    if TOKEN_PATTERN.fullmatch(value) is None:
        raise RuntimeError(f"{name} is not a valid 32-character REDCap token")
    return value


def _first_mapping(value: Any) -> dict[str, Any]:
    if isinstance(value, Mapping):
        return dict(value)
    if isinstance(value, list) and value and isinstance(value[0], Mapping):
        return dict(value[0])
    return {}


def _metadata_frame(value: Any) -> pd.DataFrame:
    frame = value.copy(deep=True) if isinstance(value, pd.DataFrame) else pd.DataFrame(value)
    if "field_name" in frame.columns:
        return frame.reset_index(drop=True)
    if frame.index.name == "field_name":
        return frame.reset_index()
    if not isinstance(frame.index, pd.RangeIndex):
        return frame.reset_index().rename(columns={"index": "field_name"})
    return frame


def _normalized_nonblank(series: pd.Series) -> pd.Series:
    return series.replace(r"^\s*$", pd.NA, regex=True)


def _stable_scalar(
    records: pd.DataFrame,
    record_id: str,
    field_name: str,
    participant_ids: pd.Index,
) -> tuple[pd.Series, pd.Series]:
    if field_name not in records.columns:
        empty = pd.Series(pd.NA, index=participant_ids, dtype="object")
        return empty, pd.Series(False, index=participant_ids, dtype=bool)
    subset = records[[record_id, field_name]].copy()
    subset[field_name] = _normalized_nonblank(subset[field_name])
    subset = subset.dropna(subset=[field_name])
    if subset.empty:
        empty = pd.Series(pd.NA, index=participant_ids, dtype="object")
        return empty, pd.Series(False, index=participant_ids, dtype=bool)
    grouped = subset.groupby(record_id, sort=False)[field_name]
    first = grouped.first().reindex(participant_ids)
    conflicts = grouped.nunique(dropna=True).gt(1).reindex(participant_ids, fill_value=False)
    conflicts = conflicts.astype(bool)
    # Never let REDCap export order decide a participant classification.
    # Conflicting values remain missing in the aggregate classification and
    # are retained as explicit review flags in the secure audit.
    return first.mask(conflicts, pd.NA), conflicts


def _choice_codes(value: Any, field_type: str) -> set[str]:
    if field_type == "yesno":
        return {"0", "1"}
    if not isinstance(value, str) or not value.strip():
        return set()
    codes: set[str] = set()
    for choice in value.split("|"):
        code, separator, _ = choice.partition(",")
        if separator and code.strip():
            codes.add(code.strip())
    return codes


def _validate_metadata_semantics(config: StudyConfig, metadata: pd.DataFrame) -> None:
    """Reject field-type or coded-choice drift when metadata exposes it."""

    if "field_name" not in metadata.columns:
        return
    indexed = metadata.set_index(metadata["field_name"].astype(str), drop=False)
    if "field_type" not in indexed.columns:
        return

    errors: list[str] = []

    def validate_field(
        field_name: str,
        allowed_types: set[str],
        configured_codes: Iterable[int] = (),
    ) -> None:
        if field_name not in indexed.index:
            return
        row = indexed.loc[field_name]
        if isinstance(row, pd.DataFrame):
            row = row.iloc[0]
        field_type = str(row.get("field_type", "")).strip().lower()
        if field_type and field_type not in allowed_types:
            errors.append(
                f"{field_name} has type {field_type!r}; expected "
                f"{'/'.join(sorted(allowed_types))}"
            )
        if not configured_codes:
            return
        choices = _choice_codes(
            row.get("select_choices_or_calculations", ""),
            field_type,
        )
        expected = {str(code) for code in configured_codes}
        if choices and not expected.issubset(choices):
            missing = ", ".join(sorted(expected - choices))
            errors.append(f"{field_name} is missing configured choice code(s): {missing}")

    validate_field(
        config.race_field,
        {"checkbox"},
        (
            *config.race_minority_codes,
            *config.race_white_codes,
            *config.race_unknown_codes,
            *config.race_hispanic_codes,
        ),
    )
    validate_field(
        config.ethnicity_field,
        {"dropdown", "radio", "yesno"},
        (*config.ethnicity_hispanic_codes, *config.ethnicity_unknown_codes),
    )
    for field_name, codes in config.secondary_ethnicity:
        validate_field(field_name, {"dropdown", "radio", "yesno"}, codes)
    for field_name in config.status_fields:
        validate_field(field_name, {"checkbox", "dropdown", "radio", "yesno"}, (1,))
    if config.dual_field:
        validate_field(config.dual_field, {"dropdown", "radio", "yesno"}, (1,))

    if errors:
        raise RuntimeError(
            f"{config.key} metadata no longer matches the configured semantics: "
            + "; ".join(errors)
        )


def _validate_export_columns(
    config: StudyConfig,
    records: pd.DataFrame,
    requested_fields: Iterable[str],
    metadata: pd.DataFrame,
) -> None:
    """Ensure requested fields were actually returned before classifying."""

    requested = set(requested_fields)
    expected_scalar = {config.record_id}
    expected_checkbox: set[str] = set()
    indexed = metadata.set_index(metadata["field_name"].astype(str), drop=False)
    for field_name in requested - {config.race_field}:
        row = indexed.loc[field_name] if field_name in indexed.index else None
        if isinstance(row, pd.DataFrame):
            row = row.iloc[0]
        field_type = (
            str(row.get("field_type", "")).strip().lower()
            if isinstance(row, pd.Series)
            else ""
        )
        if field_type == "checkbox":
            codes = _choice_codes(
                row.get("select_choices_or_calculations", ""),
                field_type,
            )
            expected_checkbox.update(f"{field_name}___{code}" for code in codes)
        else:
            expected_scalar.add(field_name)
    missing = sorted(
        (expected_scalar | expected_checkbox) - set(records.columns)
    )

    race_codes = {
        *config.race_minority_codes,
        *config.race_white_codes,
        *config.race_unknown_codes,
        *config.race_hispanic_codes,
    }
    expected_race = {f"{config.race_field}___{code}" for code in race_codes}
    missing.extend(sorted(expected_race - set(records.columns)))
    if missing:
        raise RuntimeError(
            f"{config.key} export omitted configured field column(s): "
            + ", ".join(missing)
        )


def _binary_max(
    records: pd.DataFrame,
    record_id: str,
    field_name: str,
    participant_ids: pd.Index,
) -> pd.Series:
    if field_name not in records.columns:
        return pd.Series(False, index=participant_ids, dtype=bool)
    numeric = pd.to_numeric(records[field_name], errors="coerce")
    maximum = numeric.groupby(records[record_id], sort=False).max().reindex(participant_ids)
    return maximum.eq(1).fillna(False).astype(bool)


def _binary_flag(
    records: pd.DataFrame,
    record_id: str,
    field_name: str,
    participant_ids: pd.Index,
) -> pd.Series:
    if field_name in records.columns:
        return _binary_max(records, record_id, field_name, participant_ids)
    return _checkbox_any(records, record_id, field_name, (1,), participant_ids)


def _checkbox_any(
    records: pd.DataFrame,
    record_id: str,
    field_name: str,
    codes: Sequence[int],
    participant_ids: pd.Index,
) -> pd.Series:
    columns = [
        f"{field_name}___{code}"
        for code in codes
        if f"{field_name}___{code}" in records.columns
    ]
    if not columns:
        return pd.Series(False, index=participant_ids, dtype=bool)
    numeric = records[columns].apply(pd.to_numeric, errors="coerce").fillna(0)
    any_row = numeric.max(axis=1).gt(0)
    any_participant = any_row.groupby(records[record_id], sort=False).max()
    return any_participant.reindex(participant_ids, fill_value=False).astype(bool)


def _date_candidate(
    records: pd.DataFrame,
    record_id: str,
    field_name: str,
    participant_ids: pd.Index,
) -> tuple[pd.Series, pd.Series, pd.Series]:
    if field_name not in records.columns:
        empty = pd.Series(pd.NaT, index=participant_ids, dtype="datetime64[ns]")
        flags = pd.Series(False, index=participant_ids, dtype=bool)
        return empty, flags, flags.copy()
    raw = _normalized_nonblank(records[field_name])
    parsed = pd.to_datetime(raw, errors="coerce")
    valid = parsed.notna()
    first = parsed[valid].groupby(records.loc[valid, record_id], sort=False).min()
    first = first.reindex(participant_ids)
    invalid_row = raw.notna() & parsed.isna()
    invalid = invalid_row.groupby(records[record_id], sort=False).max()
    invalid = invalid.reindex(participant_ids, fill_value=False).astype(bool)
    distinct = parsed[valid].groupby(records.loc[valid, record_id], sort=False).nunique()
    multiple = distinct.gt(1).reindex(participant_ids, fill_value=False).astype(bool)
    return first, invalid, multiple


def _reason_text(row: pd.Series, config: StudyConfig) -> str:
    reasons: list[str] = []
    for field_name in config.status_fields:
        if bool(row.get(f"flag_{field_name}", False)):
            reasons.append(f"{field_name}=Yes")
    if bool(row.get("missing_race", False)):
        reasons.append("missing race")
    if bool(row.get("race_unknown", False)):
        reasons.append("race unknown/other")
    if bool(row.get("missing_ethnicity", False)):
        reasons.append("missing/unknown ethnicity")
    for field_name in (config.ethnicity_field, *[name for name, _ in config.secondary_ethnicity]):
        if bool(row.get(f"conflict_{field_name}", False)):
            reasons.append(f"conflicting {field_name} values")
    return "; ".join(reasons) if reasons else "no review flags"


def classify_records(
    *,
    config: StudyConfig,
    records: pd.DataFrame,
    metadata_fields: set[str],
    strict_eligibility: bool = False,
) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    """Reduce a longitudinal export to one explicit audit row per participant."""

    if config.record_id not in records.columns:
        raise RuntimeError(
            f"{config.key} record export did not materialize configured ID "
            f"field {config.record_id!r}"
        )
    record_values = records[config.record_id].astype("string").str.strip()
    valid_record_id = record_values.notna() & record_values.ne("")
    records = records.loc[valid_record_id].copy()
    records[config.record_id] = record_values.loc[valid_record_id]
    participant_ids = pd.Index(pd.unique(records[config.record_id]), name=config.record_id)
    if participant_ids.empty:
        raise RuntimeError(f"{config.key} returned no participant records")

    race_codes = (
        *config.race_minority_codes,
        *config.race_white_codes,
        *config.race_unknown_codes,
        *config.race_hispanic_codes,
    )
    audit = pd.DataFrame({config.record_id: participant_ids.to_list()})
    audit.index = participant_ids

    race_any = _checkbox_any(
        records, config.record_id, config.race_field, race_codes, participant_ids
    )
    minority = _checkbox_any(
        records,
        config.record_id,
        config.race_field,
        config.race_minority_codes,
        participant_ids,
    )
    white = _checkbox_any(
        records,
        config.record_id,
        config.race_field,
        config.race_white_codes,
        participant_ids,
    )
    race_unknown = _checkbox_any(
        records,
        config.record_id,
        config.race_field,
        config.race_unknown_codes,
        participant_ids,
    )
    race_hispanic = _checkbox_any(
        records,
        config.record_id,
        config.race_field,
        config.race_hispanic_codes,
        participant_ids,
    )

    primary_ethnicity, primary_conflict = _stable_scalar(
        records, config.record_id, config.ethnicity_field, participant_ids
    )
    primary_numeric = pd.to_numeric(primary_ethnicity, errors="coerce")
    is_hispanic = primary_numeric.isin(config.ethnicity_hispanic_codes)
    explicit_unknown = primary_numeric.isin(config.ethnicity_unknown_codes)
    ethnicity_observed = primary_numeric.notna()
    ethnicity_known = primary_numeric.notna() & ~explicit_unknown

    audit[f"source_{config.ethnicity_field}_observed"] = ethnicity_observed.astype(int)
    audit[f"conflict_{config.ethnicity_field}"] = primary_conflict.astype(int)
    for field_name, hispanic_codes in config.secondary_ethnicity:
        values, conflicts = _stable_scalar(
            records, config.record_id, field_name, participant_ids
        )
        numeric = pd.to_numeric(values, errors="coerce")
        is_hispanic = is_hispanic | numeric.isin(hispanic_codes)
        ethnicity_observed = ethnicity_observed | numeric.notna()
        ethnicity_known = ethnicity_known | numeric.notna()
        audit[f"source_{field_name}_observed"] = numeric.notna().astype(int)
        audit[f"conflict_{field_name}"] = conflicts.astype(int)

    if config.race_hispanic_codes:
        is_hispanic = is_hispanic | race_hispanic
        ethnicity_observed = ethnicity_observed | race_hispanic
        ethnicity_known = ethnicity_known | race_hispanic

    audit["race_any"] = race_any.astype(int)
    audit["is_minority"] = minority.astype(int)
    audit["race_white"] = white.astype(int)
    audit["race_unknown"] = race_unknown.astype(int)
    audit["is_hispanic"] = is_hispanic.astype(int)
    audit["ethnicity_observed"] = ethnicity_observed.astype(int)
    audit["ethnicity_known"] = ethnicity_known.astype(int)
    audit["ethnicity_explicit_unknown"] = explicit_unknown.astype(int)
    audit["missing_race"] = (~race_any).astype(int)
    audit["missing_ethnicity"] = (~ethnicity_known).astype(int)

    for field_name in config.status_fields:
        audit[f"flag_{field_name}"] = _binary_flag(
            records, config.record_id, field_name, participant_ids
        ).astype(int)
    if config.dual_field:
        audit["dual_enrolled"] = _binary_flag(
            records, config.record_id, config.dual_field, participant_ids
        ).astype(int)

    date_coverage: list[dict[str, Any]] = []
    for field_name, field_label in config.date_candidates:
        if field_name not in metadata_fields:
            date_coverage.append(
                {
                    "Project": config.key,
                    "Field": field_name,
                    "Description": field_label,
                    "Participants with valid date": 0,
                    "Participants with invalid date": 0,
                    "Participants with multiple dates": 0,
                    "Interpretation": "Field not present; not used as enrollment date",
                }
            )
            continue
        first, invalid, multiple = _date_candidate(
            records, config.record_id, field_name, participant_ids
        )
        audit[f"first_{field_name}"] = first.dt.strftime("%Y-%m-%d").fillna("")
        audit[f"invalid_{field_name}"] = invalid.astype(int)
        audit[f"multiple_{field_name}"] = multiple.astype(int)
        date_coverage.append(
            {
                "Project": config.key,
                "Field": field_name,
                "Description": field_label,
                "Participants with valid date": int(first.notna().sum()),
                "Participants with invalid date": int(invalid.sum()),
                "Participants with multiple dates": int(multiple.sum()),
                "Interpretation": "Diagnostic candidate; not used as enrollment date",
            }
        )

    exclusion_hit = pd.Series(False, index=participant_ids)
    for field_name in config.exclusion_fields:
        exclusion_hit = exclusion_hit | audit[f"flag_{field_name}"].astype(bool)
    review_hit = pd.Series(False, index=participant_ids)
    for field_name in config.review_fields:
        review_hit = review_hit | audit[f"flag_{field_name}"].astype(bool)
    conflict_columns = [
        column for column in audit.columns if column.startswith("conflict_")
    ]
    conflict_hit = (
        audit[conflict_columns].astype(bool).any(axis=1)
        if conflict_columns
        else pd.Series(False, index=participant_ids)
    )
    issue_hit = (
        exclusion_hit
        | review_hit
        | audit["missing_race"].astype(bool)
        | audit["race_unknown"].astype(bool)
        | audit["missing_ethnicity"].astype(bool)
        | conflict_hit
    )
    excluded = exclusion_hit if strict_eligibility else pd.Series(False, index=participant_ids)
    audit["decision"] = "included"
    audit.loc[issue_hit & ~excluded, "decision"] = "flagged-review"
    audit.loc[excluded, "decision"] = "excluded"
    audit["in_cumulative"] = (~excluded).astype(int)
    audit["reason"] = audit.apply(lambda row: _reason_text(row, config), axis=1)
    audit = audit.reset_index(drop=True)
    return audit, date_coverage


def _project_factory(
    api_url: str,
    token: str,
    *,
    timeout: tuple[int, int] = (10, 90),
) -> Project:
    return Project(api_url, token, timeout=timeout)


def fetch_study(
    *,
    config: StudyConfig,
    token: str,
    api_url: str,
    strict_eligibility: bool,
    project_factory: Callable[..., Any] = _project_factory,
) -> StudyResult:
    """Fetch and classify one study using export-only REDCap operations."""

    try:
        project = project_factory(api_url, token, timeout=(10, 90))
        info = _first_mapping(project.export_project_info(format_type="json"))
        actual_project_id = int(info.get("project_id"))
        if actual_project_id != config.expected_project_id:
            raise RuntimeError(
                f"{config.key} token resolved to unexpected REDCap project "
                f"{actual_project_id}; expected {config.expected_project_id}"
            )

        metadata = _metadata_frame(project.export_metadata(format_type="df"))
        if "field_name" not in metadata.columns:
            raise RuntimeError(f"{config.key} metadata has no field_name column")
        metadata_fields = set(metadata["field_name"].astype(str))
        required_fields = {
            config.record_id,
            config.race_field,
            config.ethnicity_field,
            *config.status_fields,
            *[name for name, _ in config.secondary_ethnicity],
        }
        if config.dual_field:
            required_fields.add(config.dual_field)
        missing_required = sorted(required_fields - metadata_fields)
        if missing_required:
            raise RuntimeError(
                f"{config.key} is missing required fields: {', '.join(missing_required)}"
            )
        _validate_metadata_semantics(config, metadata)

        requested = [config.race_field, config.ethnicity_field]
        requested.extend(name for name, _ in config.secondary_ethnicity)
        requested.extend(name for name in config.status_fields if name in metadata_fields)
        if config.dual_field and config.dual_field in metadata_fields:
            requested.append(config.dual_field)
        requested.extend(
            name for name, _ in config.date_candidates if name in metadata_fields
        )
        requested = list(dict.fromkeys(requested))
        records_value = project.export_records(
            format_type="df",
            fields=requested,
            raw_or_label="raw",
            df_kwargs={"dtype": str, "keep_default_na": False},
        )
        records = (
            records_value.copy(deep=True)
            if isinstance(records_value, pd.DataFrame)
            else pd.DataFrame(records_value)
        )
        try:
            records = records.reset_index()
        except ValueError:
            records = records.reset_index(drop=True)
        _validate_export_columns(config, records, requested, metadata)
        audit, date_coverage = classify_records(
            config=config,
            records=records,
            metadata_fields=metadata_fields,
            strict_eligibility=strict_eligibility,
        )
    except Exception as exc:
        detail = sanitize_error(exc)
        raise RuntimeError(f"{config.key} REDCap export failed: {detail}") from exc

    counted = audit["in_cumulative"].eq(1)
    strict_exclusion = pd.Series(False, index=audit.index)
    for field_name in config.exclusion_fields:
        strict_exclusion = strict_exclusion | audit[f"flag_{field_name}"].eq(1)
    strict_counted = ~strict_exclusion
    live_counts = {
        "Total": int(counted.sum()),
        "Minority": int((counted & audit["is_minority"].eq(1)).sum()),
        "Hispanic": int((counted & audit["is_hispanic"].eq(1)).sum()),
    }
    data_quality = {
        "Unique participants": int(len(audit)),
        "Included without flags": int(audit["decision"].eq("included").sum()),
        "Flagged for review": int(audit["decision"].eq("flagged-review").sum()),
        "Excluded by strict policy": int(audit["decision"].eq("excluded").sum()),
        "Counted in cumulative": int(counted.sum()),
        "Would be excluded under strict policy": int(strict_exclusion.sum()),
        "Strict-policy sensitivity total": int(strict_counted.sum()),
        "Strict-policy sensitivity minority": int(
            (strict_counted & audit["is_minority"].eq(1)).sum()
        ),
        "Strict-policy sensitivity Hispanic/Latino": int(
            (strict_counted & audit["is_hispanic"].eq(1)).sum()
        ),
        "Race populated": int(audit["race_any"].sum()),
        "Race missing": int(audit["missing_race"].sum()),
        "Race unknown/other": int(audit["race_unknown"].sum()),
        "Ethnicity observed": int(audit["ethnicity_observed"].sum()),
        "Ethnicity known": int(audit["ethnicity_known"].sum()),
        "Ethnicity missing/unknown": int(audit["missing_ethnicity"].sum()),
        "Racial minority": int(audit["is_minority"].sum()),
        "Hispanic/Latino": int(audit["is_hispanic"].sum()),
        "Dual-enrolled in NANO": int(audit.get("dual_enrolled", pd.Series(dtype=int)).sum()),
    }
    qa_checks = [
        {
            "Project": config.key,
            "Check": "Configured REDCap project identity",
            "Passed": actual_project_id == config.expected_project_id,
            "Detail": f"PID {actual_project_id}",
        },
        {
            "Project": config.key,
            "Check": "One audit row per participant",
            "Passed": bool(audit[config.record_id].is_unique),
            "Detail": "Audit keys are unique",
        },
        {
            "Project": config.key,
            "Check": "Cumulative total reconciles to audit",
            "Passed": live_counts["Total"] == int(audit["in_cumulative"].sum()),
            "Detail": "Matches calculated live count",
        },
        {
            "Project": config.key,
            "Check": "Demographic categories do not exceed total",
            "Passed": (
                live_counts["Minority"] <= live_counts["Total"]
                and live_counts["Hispanic"] <= live_counts["Total"]
            ),
            "Detail": "Both category counts are within the live total",
        },
        {
            "Project": config.key,
            "Check": "No unverified enrollment date inferred",
            "Passed": True,
            "Detail": "All date fields are diagnostic candidates only",
        },
    ]
    return StudyResult(
        config=config,
        project_info=info,
        metadata_fields=metadata_fields,
        participant_audit=audit,
        date_coverage=date_coverage,
        live_counts=live_counts,
        data_quality=data_quality,
        qa_checks=qa_checks,
        raw_row_count=int(len(records)),
        strict_eligibility=strict_eligibility,
    )


def build_milestones(config: StudyConfig) -> list[date]:
    milestones: list[date] = []
    for year in range(config.milestone_start.year, config.milestone_end.year + 1):
        for month in config.milestone_months:
            candidate = date(year, month, 1)
            if config.milestone_start <= candidate <= config.milestone_end:
                milestones.append(candidate)
    return milestones


def _pad(values: Sequence[Any], length: int) -> list[Any]:
    return [values[index] if index < len(values) else None for index in range(length)]


def _current_milestone_index(milestones: Sequence[date], report_date: date) -> int:
    return next(
        (index for index, milestone in enumerate(milestones) if milestone >= report_date),
        len(milestones) - 1,
    )


def report_rows(result: StudyResult, report_date: date) -> tuple[list[date], list[dict[str, Any]]]:
    config = result.config
    milestones = build_milestones(config)
    current_index = _current_milestone_index(milestones, report_date)
    rows: list[dict[str, Any]] = []
    for category in CATEGORIES:
        previous = _pad(config.previous_targets.get(category, ()), len(milestones))
        current = _pad(config.current_targets.get(category, ()), len(milestones))
        actual = _pad(config.historical_actuals.get(category, ()), len(milestones))
        rows.extend(
            [
                {
                    "category": category,
                    "kind": "previous",
                    "label": f"Previous Target: {CATEGORY_LABELS[category]}",
                    "values": previous,
                    "live": previous[current_index],
                },
                {
                    "category": category,
                    "kind": "current",
                    "label": f"Current Target: {CATEGORY_LABELS[category]}",
                    "values": current,
                    "live": current[current_index],
                },
                {
                    "category": category,
                    "kind": "actual",
                    "label": (
                        "Published Historical Actual / Live-as-of Count: "
                        f"{CATEGORY_LABELS[category]}"
                    ),
                    "values": actual,
                    "live": result.live_counts[category],
                },
                {
                    "category": category,
                    "kind": "ratio",
                    "label": f"Actual / Current Target: {CATEGORY_LABELS[category]}",
                    "values": [None] * len(milestones),
                    "live": None,
                },
                {
                    "category": category,
                    "kind": "status",
                    "label": f"Status: {CATEGORY_LABELS[category]}",
                    "values": [None] * len(milestones),
                    "live": None,
                },
            ]
        )
    return milestones, rows


def _ratio(actual: Any, target: Any) -> float | None:
    if actual is None or isinstance(actual, str) or target in (None, 0):
        return None
    return float(actual) / float(target)


def _public_ratio(actual: Any, target: Any) -> float | None:
    if isinstance(_public_count(actual), str):
        return None
    return _ratio(actual, target)


def _status(actual: Any, target: Any) -> str:
    if actual is None:
        return ""
    if isinstance(actual, str):
        return "N/A"
    if target in (None, 0):
        return "N/A"
    return "On Target" if float(actual) >= float(target) else "Below Target"


def _public_status(actual: Any, target: Any) -> str:
    if isinstance(_public_count(actual), str):
        return "Suppressed"
    return _status(actual, target)


def _public_count(value: Any) -> Any:
    """Suppress non-zero participant-derived cells below the public threshold."""

    if isinstance(value, bool):
        return value
    if isinstance(value, int) and 0 < value < PUBLIC_SMALL_CELL_THRESHOLD:
        return f"<{PUBLIC_SMALL_CELL_THRESHOLD}"
    return value


def _public_banded_count(value: Any) -> Any:
    """Publish diagnostic counts in ten-person bands to prevent subtraction."""

    if isinstance(value, bool) or not isinstance(value, int):
        return value
    if value == 0:
        return 0
    if value < PUBLIC_SMALL_CELL_THRESHOLD:
        return f"<{PUBLIC_SMALL_CELL_THRESHOLD}"
    lower = value - (value % PUBLIC_SMALL_CELL_THRESHOLD)
    upper = lower + PUBLIC_SMALL_CELL_THRESHOLD - 1
    return f"{lower}\N{EN DASH}{upper}"


def _public_data_quality(result: StudyResult) -> dict[str, Any]:
    # Complementary "populated"/"included" values and strict-policy totals can
    # disclose suppressed cells by subtraction. Exact values remain available
    # in the ignored participant audit; public diagnostics use ten-person bands.
    public_metrics = (
        "Flagged for review",
        "Would be excluded under strict policy",
        "Race missing",
        "Race unknown/other",
        "Ethnicity missing/unknown",
        "Dual-enrolled in NANO",
    )
    return {
        metric: _public_banded_count(result.data_quality[metric])
        for metric in public_metrics
    }


def _public_date_coverage(
    rows: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    count_fields = {
        "Participants with valid date",
        "Participants with invalid date",
        "Participants with multiple dates",
    }
    return [
        {
            key: _public_banded_count(value) if key in count_fields else value
            for key, value in row.items()
        }
        for row in rows
    ]


def _html_value(kind: str, value: Any) -> str:
    if value is None:
        return ""
    if kind == "ratio":
        return f"{float(value):.0%}"
    return escape(str(value))


def render_html(result: StudyResult, report_date: date, generated_at: datetime) -> str:
    config = result.config
    milestones, rows = report_rows(result, report_date)
    by_category = {
        category: {row["kind"]: row for row in rows if row["category"] == category}
        for category in CATEGORIES
    }
    header_cells = "".join(
        f"<th>{escape(MONTH_LABELS[m.month])}<small>{m.year}</small></th>"
        for m in milestones
    )
    body: list[str] = []
    for category in CATEGORIES:
        category_rows = by_category[category]
        for kind in ("previous", "current", "actual", "ratio", "status"):
            row = category_rows[kind]
            values: list[Any]
            if kind == "ratio":
                if config.targets_state == "verified":
                    values = [
                        _public_ratio(actual, target)
                        for actual, target in zip(
                            category_rows["actual"]["values"],
                            category_rows["current"]["values"],
                        )
                    ]
                    live = _public_ratio(
                        category_rows["actual"]["live"],
                        category_rows["current"]["live"],
                    )
                else:
                    values = [None] * len(milestones)
                    live = None
            elif kind == "status":
                if config.targets_state == "verified":
                    values = [
                        _public_status(actual, target)
                        for actual, target in zip(
                            category_rows["actual"]["values"],
                            category_rows["current"]["values"],
                        )
                    ]
                    live = _public_status(
                        category_rows["actual"]["live"],
                        category_rows["current"]["live"],
                    )
                else:
                    values = ["N/A — target unverified"] * len(milestones)
                    live = "N/A — target unverified"
            else:
                values = row["values"]
                live = row["live"]
            if kind == "actual":
                values = [_public_count(value) for value in values]
                live = _public_count(live)
            cells = "".join(
                f'<td class="{escape(kind)}">{_html_value(kind, value)}</td>'
                for value in values
            )
            body.append(
                f'<tr class="{escape(kind)}"><th>{escape(row["label"])}</th>'
                f'{cells}<td class="live {_html_value(kind, live).replace(" ", "-").lower()}">'
                f"{_html_value(kind, live)}</td></tr>"
            )

    public_quality = _public_data_quality(result)
    quality_rows = "".join(
        f"<tr><th>{escape(metric)}</th><td>{escape(str(value))}</td></tr>"
        for metric, value in public_quality.items()
    )
    public_dates = _public_date_coverage(result.date_coverage)
    date_rows = "".join(
        "<tr>"
        f"<td>{escape(str(item['Field']))}</td>"
        f"<td>{escape(str(item['Description']))}</td>"
        f"<td>{escape(str(item['Participants with valid date']))}</td>"
        f"<td>{escape(str(item['Participants with invalid date']))}</td>"
        f"<td>{escape(str(item['Participants with multiple dates']))}</td>"
        f"<td>{escape(str(item['Interpretation']))}</td>"
        "</tr>"
        for item in public_dates
    )
    policy = (
        "Strict eligibility is active; records flagged ineligible or unenrolled are excluded."
        if result.strict_eligibility
        else "All unique study records are counted; ineligible/unenrolled and data-quality "
        "issues remain counted but are explicitly flagged for review."
    )
    live_total_label = (
        "Strict-policy live total"
        if result.strict_eligibility
        else "All-record live total (policy under review)"
    )
    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{escape(config.label)} recruitment report</title>
<style>
:root{{--navy:#17365d;--blue:#4f81bd;--light:#eef3f8;--orange:#f6a821;--ink:#1f2933;}}
*{{box-sizing:border-box}} body{{margin:0;background:#f5f7fa;color:var(--ink);
font:14px/1.45 "Libre Franklin","Segoe UI",Arial,sans-serif}}
main{{max-width:1500px;margin:0 auto;padding:28px}} h1{{margin:0 0 6px;font-size:28px}}
.subtitle{{color:#52606d;margin-bottom:18px}} .banner{{padding:12px 14px;margin:10px 0;
border-left:5px solid var(--orange);background:#fff7e6;border-radius:4px}}
.cards{{display:grid;grid-template-columns:repeat(3,minmax(180px,1fr));gap:12px;margin:18px 0}}
.card{{background:white;padding:16px;border-radius:8px;box-shadow:0 1px 4px #0001}}
.card strong{{display:block;font-size:30px;color:var(--navy)}} .card span{{color:#52606d}}
.scroll{{overflow:auto;background:white;border-radius:8px;box-shadow:0 1px 4px #0001}}
table{{border-collapse:separate;border-spacing:0;width:max-content;min-width:100%}}
th,td{{padding:8px 10px;border-right:1px solid #d9e2ec;border-bottom:1px solid #d9e2ec;
text-align:center;white-space:nowrap}} thead th{{background:var(--blue);color:white;position:sticky;top:0}}
thead small{{display:block;font-weight:400}} tbody th{{text-align:left;position:sticky;left:0;
background:white;min-width:355px;z-index:1}} td.live,thead th.live{{background:#fff1d6;font-weight:700}}
tr.actual th,tr.current th{{font-weight:700}} tr.ratio td{{font-variant-numeric:tabular-nums}}
.section{{margin-top:24px;background:white;border-radius:8px;padding:18px;box-shadow:0 1px 4px #0001}}
.section table{{width:100%;min-width:0}} .section tbody th{{position:static;min-width:0}}
.section td,.section th{{white-space:normal;text-align:left}} footer{{color:#7b8794;margin-top:18px}}
@media(max-width:760px){{main{{padding:14px}}.cards{{grid-template-columns:1fr}}}}
</style>
</head>
<body><main>
<h1>{escape(config.label)} recruitment milestones</h1>
<div class="subtitle">REDCap PID {config.expected_project_id} · report date
{report_date.isoformat()} · generated {generated_at.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")}</div>
<div class="banner"><strong>Counting policy:</strong> {escape(policy)}</div>
<div class="banner"><strong>Date policy:</strong> No date candidate is treated as a verified
protocol-enrollment date. Published history is preserved; current live totals appear only in
the separate as-of column.</div>
<div class="banner"><strong>Targets:</strong> {escape(config.targets_state)} —
{escape(config.targets_source)}.</div>
<div class="banner"><strong>Demographics:</strong> {escape(config.demographics_state)} —
{escape(config.demographics_note)}.</div>
<div class="cards">
<div class="card"><strong>{escape(str(_public_count(result.live_counts['Total'])))}</strong><span>{escape(live_total_label)} as of {report_date}</span></div>
<div class="card"><strong>{escape(str(_public_count(result.live_counts['Minority'])))}</strong><span>Racial minority ({escape(config.demographics_state)})</span></div>
<div class="card"><strong>{escape(str(_public_count(result.live_counts['Hispanic'])))}</strong><span>Observed Hispanic/Latino ({escape(config.demographics_state)}); {escape(str(public_quality['Ethnicity missing/unknown']))} missing/unknown</span></div>
</div>
<div class="scroll"><table>
<thead><tr><th>Metric</th>{header_cells}<th class="live">Live as of<small>{report_date}</small></th></tr></thead>
<tbody>{''.join(body)}</tbody></table></div>
<section class="section"><h2>Aggregate data quality</h2><table><tbody>{quality_rows}</tbody></table></section>
<section class="section"><h2>Date-field coverage</h2>
<p>Coverage is diagnostic only and does not establish enrollment semantics.</p>
<table><thead><tr><th>Field</th><th>Description</th><th>Valid participants</th>
<th>Invalid</th><th>Multiple values</th><th>Interpretation</th></tr></thead>
<tbody>{date_rows}</tbody></table></section>
<footer>Participant-level audit rows are stored only in the ignored secure output directory.
Participant-derived cells from 1 through 9 are shown as &lt;10; diagnostic counts and
temporally complementary cumulative cells are banded. Apply IRB and data-use rules
before sharing.</footer>
</main></body></html>"""


def _protect_csv_value(value: Any) -> Any:
    if isinstance(value, str) and value.lstrip().startswith(FORMULA_PREFIXES):
        return "'" + value
    return value


def _write_secure_audit(path: Path, frame: pd.DataFrame) -> None:
    safe = frame.copy(deep=True)
    for column in safe.columns:
        safe[column] = safe[column].map(_protect_csv_value)
    descriptor, temporary_name = tempfile.mkstemp(
        dir=path.parent,
        prefix=f".{path.name}.",
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        safe.to_csv(temporary_path, index=False)
        temporary_path.chmod(0o600)
        os.replace(temporary_path, path)
    finally:
        temporary_path.unlink(missing_ok=True)


def _write_secure_text(path: Path, text: str) -> None:
    descriptor, temporary_name = tempfile.mkstemp(
        dir=path.parent,
        prefix=f".{path.name}.",
        text=True,
    )
    temporary_path = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8") as handle:
            handle.write(text)
        temporary_path.chmod(0o600)
        os.replace(temporary_path, path)
    finally:
        temporary_path.unlink(missing_ok=True)


def _style_header(ws: Worksheet, row: int, start: int, end: int) -> None:
    for cell in ws.iter_cols(min_col=start, max_col=end, min_row=row, max_row=row):
        item = cell[0]
        item.fill = PatternFill("solid", fgColor=BLUE)
        item.font = Font(color=WHITE, bold=True)
        item.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_tabular_sheet(ws: Worksheet) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    thin = Side(style="thin", color="D9E2EC")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        longest = max((len(str(cell.value or "")) for cell in column_cells), default=8)
        ws.column_dimensions[letter].width = min(max(longest + 2, 12), 55)
    ws.auto_filter.ref = ws.dimensions


def _append_table_sheet(
    workbook: Workbook,
    name: str,
    rows: Sequence[Mapping[str, Any]],
) -> Worksheet:
    ws = workbook.create_sheet(name[:31])
    if not rows:
        ws.append(["No rows"])
        return ws
    columns = list(rows[0].keys())
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])
    _style_header(ws, 1, 1, len(columns))
    _style_tabular_sheet(ws)
    return ws


def _write_study_sheet(
    workbook: Workbook,
    result: StudyResult,
    report_date: date,
    counting_policy: str,
) -> Worksheet:
    ws = workbook.create_sheet(result.config.key)
    milestones, rows = report_rows(result, report_date)
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{result.config.label} Recruitment Milestones"
    ws["A1"].font = Font(size=16, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(milestones) + 2)
    ws["A2"] = (
        f"Report date {report_date.isoformat()}\n"
        f"Counting policy: {counting_policy}\n"
        "No date candidate is treated as verified protocol enrollment; live counts "
        "are not backfilled into a milestone."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws["A2"].font = Font(size=9)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(milestones) + 2)
    headers = ["Metric", *[f"{MONTH_LABELS[item.month]} {item.year}" for item in milestones], f"Live as of {report_date}"]
    ws.append([])
    ws.append(headers)
    _style_header(ws, 4, 1, len(headers))
    ws.cell(row=4, column=len(headers)).fill = PatternFill("solid", fgColor=ORANGE)

    by_category: dict[str, dict[str, int]] = {}
    for row in rows:
        output_row = ws.max_row + 1
        ws.cell(output_row, 1, row["label"])
        kind = row["kind"]
        category = row["category"]
        by_category.setdefault(category, {})[kind] = output_row
        if kind not in {"ratio", "status"}:
            values = row["values"]
            live = row["live"]
            if kind == "actual":
                values = [_public_count(value) for value in values]
                live = _public_count(live)
            for column, value in enumerate(values, start=2):
                ws.cell(output_row, column, value)
            ws.cell(output_row, len(headers), live)
        if kind == "previous":
            fill = PatternFill("solid", fgColor=LIGHT_GREY)
        elif kind == "current":
            fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        elif kind == "actual":
            fill = PatternFill("solid", fgColor=WHITE)
        else:
            fill = PatternFill("solid", fgColor="F8FAFC")
        for cell in ws[output_row]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(output_row, len(headers)).fill = PatternFill("solid", fgColor=LIGHT_ORANGE)

    for category, row_map in by_category.items():
        for column in range(2, len(headers) + 1):
            actual_row = row_map["actual"]
            current_row = row_map["current"]
            ratio_row = row_map["ratio"]
            status_row = row_map["status"]
            actual = ws.cell(actual_row, column).value
            current = ws.cell(current_row, column).value
            if result.config.targets_state == "verified":
                ratio_value = _ratio(actual, current)
                status_value = _status(actual, current)
            else:
                ratio_value = None
                status_value = "N/A — target unverified"
            ws.cell(ratio_row, column, ratio_value)
            ws.cell(ratio_row, column).number_format = "0%"
            ws.cell(status_row, column, status_value)

    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 55
    for column in range(2, len(headers)):
        ws.column_dimensions[get_column_letter(column)].width = 14
    ws.column_dimensions[get_column_letter(len(headers))].width = 20
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 48
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{ws.max_row}"
    thin = Side(style="thin", color="D9E2EC")
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = Border(bottom=thin)
    return ws


def write_workbook(
    path: Path,
    studies: Sequence[StudyResult],
    report_date: date,
    generated_at: datetime,
    counting_policy: str,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for study in studies:
        _write_study_sheet(workbook, study, report_date, counting_policy)

    summary_rows: list[dict[str, Any]] = []
    quality_rows: list[dict[str, Any]] = []
    date_rows: list[dict[str, Any]] = []
    provenance_rows: list[dict[str, Any]] = []
    qa_rows: list[dict[str, Any]] = []
    for study in studies:
        summary_rows.append(
            {
                "Project": study.config.key,
                "REDCap PID": study.project_id,
                "Project title": study.project_info.get("project_title", study.config.label),
                "Report date": report_date,
                "Generated UTC": generated_at.replace(tzinfo=None),
                "Counting policy": counting_policy,
                "Public count policy": (
                    "Participant-derived cells 1–9 are <10; diagnostics and "
                    "temporally complementary cumulative cells are banded"
                ),
                "Live total": _public_count(study.live_counts["Total"]),
                "Live racial minority": _public_count(study.live_counts["Minority"]),
                "Observed live Hispanic/Latino": _public_count(
                    study.live_counts["Hispanic"]
                ),
                "Ethnicity missing/unknown": _public_data_quality(study)[
                    "Ethnicity missing/unknown"
                ],
                "Demographic classification state": study.config.demographics_state,
                "Demographic classification note": study.config.demographics_note,
            }
        )
        for metric, value in _public_data_quality(study).items():
            quality_rows.append(
                {"Project": study.config.key, "Metric": metric, "Value": value}
            )
        date_rows.extend(_public_date_coverage(study.date_coverage))
        for category in CATEGORIES:
            provenance_rows.append(
                {
                    "Project": study.config.key,
                    "Metric": CATEGORY_LABELS[category],
                    "Target state": study.config.targets_state,
                    "Target source": study.config.targets_source,
                    "Historical source": (
                        "Published cumulative history"
                        if study.config.historical_actuals.get(category)
                        else "Not supplied"
                    ),
                    "Enrollment date status": "No verified field; candidates are diagnostic only",
                    "Demographic classification state": study.config.demographics_state,
                    "Demographic classification note": study.config.demographics_note,
                    "Ratio definition": (
                        "Calculated only after target verification; otherwise suppressed"
                    ),
                }
            )
        qa_rows.extend(study.qa_checks)

    for name, rows in (
        ("Run Summary", summary_rows),
        ("Data Quality", quality_rows),
        ("Date Candidates", date_rows),
        ("Provenance", provenance_rows),
        ("QA Checks", qa_rows),
    ):
        ws = _append_table_sheet(workbook, name, rows)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, date):
                    cell.number_format = "yyyy-mm-dd"
                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm"
        if name == "QA Checks":
            for row in range(2, ws.max_row + 1):
                passed = bool(ws.cell(row, 3).value)
                ws.cell(row, 3).fill = PatternFill(
                    "solid", fgColor=GREEN if passed else RED
                )
    workbook.save(path)


def _manifest(
    result: ReportResult,
    api_url: str,
) -> dict[str, Any]:
    projects: dict[str, Any] = {}
    for key, study in result.studies.items():
        projects[key] = {
            "project_id": study.project_id,
            "project_title": study.project_info.get("project_title", study.config.label),
            "raw_exported_rows": study.raw_row_count,
            "live_counts": {
                metric: _public_count(value)
                for metric, value in study.live_counts.items()
            },
            "data_quality": _public_data_quality(study),
            "date_candidates": _public_date_coverage(study.date_coverage),
            "targets_state": study.config.targets_state,
            "targets_source": study.config.targets_source,
            "demographics_state": study.config.demographics_state,
            "demographics_note": study.config.demographics_note,
            "qa_checks": study.qa_checks,
        }
    return {
        "schema_version": 1,
        "generated_at": result.generated_at.isoformat(),
        "report_date": result.report_date.isoformat(),
        "api_url": api_url,
        "counting_policy": result.counting_policy,
        "public_count_policy": (
            "Participant-derived cells from 1 through 9 are shown as <10; "
            "diagnostic and temporally complementary cumulative counts are banded."
        ),
        "date_policy": (
            "No candidate field is treated as verified protocol enrollment; "
            "current live counts are kept in a separate as-of column."
        ),
        "projects": projects,
    }


def _secure_summary(result: ReportResult, api_url: str) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "classification": "restricted-local",
        "generated_at": result.generated_at.isoformat(),
        "report_date": result.report_date.isoformat(),
        "api_url": api_url,
        "counting_policy": result.counting_policy,
        "projects": {
            key: {
                "project_id": study.project_id,
                "project_title": study.project_info.get(
                    "project_title",
                    study.config.label,
                ),
                "raw_exported_rows": study.raw_row_count,
                "live_counts": study.live_counts,
                "data_quality": study.data_quality,
                "date_candidates": study.date_coverage,
                "qa_checks": study.qa_checks,
            }
            for key, study in result.studies.items()
        },
    }


def generate_reports(
    *,
    report_date: date | None = None,
    public_dir: str | Path = "recruitment_outputs",
    secure_dir: str | Path = "recruitment_audit_secure",
    legacy_nano_dir: str | Path = "nano_recruitment_outputs",
    env_file: str | Path = ".env",
    strict_eligibility: bool = False,
    api_url: str | None = None,
    project_factory: Callable[..., Any] = _project_factory,
) -> ReportResult:
    """Fetch both studies and refresh every canonical and compatibility output."""

    load_env_file(env_file)
    resolved_date = report_date or date.today()
    generated_at = datetime.now(timezone.utc)
    resolved_api_url = (api_url or os.environ.get("REDCAP_API_URL") or DEFAULT_API_URL).strip()
    tokens = {
        "NANO": _required_token("NANO_API_TOKEN"),
        "NICO": _required_token("NICO_API_TOKEN"),
    }
    counting_policy = (
        "strict eligibility: ineligible/unenrolled records excluded"
        if strict_eligibility
        else "all unique records counted; ineligible/unenrolled and quality issues flagged"
    )
    studies: dict[str, StudyResult] = {}
    for config in STUDIES:
        studies[config.key] = fetch_study(
            config=config,
            token=tokens[config.key],
            api_url=resolved_api_url,
            strict_eligibility=strict_eligibility,
            project_factory=project_factory,
        )
    result = ReportResult(
        report_date=resolved_date,
        generated_at=generated_at,
        counting_policy=counting_policy,
        studies=studies,
    )

    public_path = Path(public_dir)
    secure_path = Path(secure_dir)
    legacy_path = Path(legacy_nano_dir)
    public_path.mkdir(parents=True, exist_ok=True)
    secure_path.mkdir(parents=True, exist_ok=True)
    legacy_path.mkdir(parents=True, exist_ok=True)
    secure_path.chmod(0o700)
    stamp = resolved_date.isoformat()

    for key, study in studies.items():
        html_path = public_path / f"{key.lower()}_recruitment_milestones_{stamp}.html"
        html_path.write_text(
            render_html(study, resolved_date, generated_at),
            encoding="utf-8",
        )
        result.paths[f"{key.lower()}_html"] = html_path
        audit_path = secure_path / f"{key.lower()}_participant_audit_{stamp}.csv"
        _write_secure_audit(audit_path, study.participant_audit)
        result.paths[f"{key.lower()}_audit"] = audit_path

    secure_summary_path = secure_path / f"recruitment_run_summary_{stamp}.json"
    _write_secure_text(
        secure_summary_path,
        json.dumps(
            _secure_summary(result, resolved_api_url),
            indent=2,
            ensure_ascii=False,
        )
        + "\n",
    )
    result.paths["secure_summary"] = secure_summary_path

    workbook_path = public_path / f"recruitment_milestones_{stamp}.xlsx"
    write_workbook(
        workbook_path,
        list(studies.values()),
        resolved_date,
        generated_at,
        counting_policy,
    )
    result.paths["workbook"] = workbook_path

    manifest_path = public_path / f"recruitment_run_manifest_{stamp}.json"
    manifest_path.write_text(
        json.dumps(_manifest(result, resolved_api_url), indent=2, ensure_ascii=False)
        + "\n",
        encoding="utf-8",
    )
    result.paths["manifest"] = manifest_path

    nano = studies["NANO"]
    legacy_html = legacy_path / "nano_recruitment_milestones.html"
    legacy_html.write_text(
        render_html(nano, resolved_date, generated_at),
        encoding="utf-8",
    )
    legacy_workbook = legacy_path / "nano_recruitment_milestones.xlsx"
    write_workbook(
        legacy_workbook,
        [nano],
        resolved_date,
        generated_at,
        counting_policy,
    )
    result.paths["legacy_nano_html"] = legacy_html
    result.paths["legacy_nano_workbook"] = legacy_workbook

    for key, study in studies.items():
        counts = study.live_counts
        print(
            f"[{key}] PID {study.project_id}: "
            f"total={_public_count(counts['Total'])}, "
            f"minority={_public_count(counts['Minority'])}, "
            f"Hispanic={_public_count(counts['Hispanic'])}"
        )
    print(f"Wrote aggregate outputs to {public_path}")
    print(f"Wrote restricted audits to {secure_path}")
    return result


def _parse_report_date(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError("report date must be YYYY-MM-DD") from exc


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Refresh aggregate NANO/NICO recruitment reports and secure audits."
    )
    parser.add_argument("--report-date", type=_parse_report_date, default=None)
    parser.add_argument("--public-dir", default="recruitment_outputs")
    parser.add_argument("--secure-dir", default="recruitment_audit_secure")
    parser.add_argument("--legacy-nano-dir", default="nano_recruitment_outputs")
    parser.add_argument("--env-file", default=".env")
    parser.add_argument(
        "--strict-eligibility",
        action="store_true",
        help="Exclude records flagged ineligible or unenrolled from live counts.",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        generate_reports(
            report_date=args.report_date,
            public_dir=args.public_dir,
            secure_dir=args.secure_dir,
            legacy_nano_dir=args.legacy_nano_dir,
            env_file=args.env_file,
            strict_eligibility=args.strict_eligibility,
        )
    except Exception as exc:
        print(f"Recruitment report failed: {sanitize_error(exc)}")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
