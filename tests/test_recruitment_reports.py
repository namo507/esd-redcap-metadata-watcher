from __future__ import annotations

from datetime import date
import json
import os
from pathlib import Path
import stat
import traceback

from openpyxl import load_workbook
import pandas as pd
import pytest

from recruitment_reports import (
    STUDIES,
    _public_ratio,
    _public_status,
    classify_records,
    generate_reports,
    load_env_file,
    report_rows,
)


NANO_TOKEN = "A" * 32
NICO_TOKEN = "B" * 32


def _metadata(config, extra_fields=()):
    names = {
        config.record_id,
        config.race_field,
        config.ethnicity_field,
        *config.status_fields,
        *[field for field, _ in config.secondary_ethnicity],
        *[field for field, _ in config.date_candidates],
        *extra_fields,
    }
    if config.dual_field:
        names.add(config.dual_field)
    return pd.DataFrame(
        {
            "field_name": sorted(names),
            "field_label": sorted(names),
            "form_name": "test",
        }
    ).set_index("field_name")


class FakeProject:
    def __init__(self, config, records, *, project_id=None):
        self.config = config
        self.records = records
        self._project_id = project_id or config.expected_project_id

    def export_project_info(self, *, format_type):
        assert format_type == "json"
        return {
            "project_id": self._project_id,
            "project_title": self.config.label,
            "is_longitudinal": 1,
        }

    def export_metadata(self, *, format_type):
        assert format_type == "df"
        return _metadata(self.config)

    def export_records(self, *, format_type, fields, raw_or_label, df_kwargs):
        assert format_type == "df"
        assert raw_or_label == "raw"
        assert self.config.record_id not in fields
        assert df_kwargs == {"dtype": str, "keep_default_na": False}
        return self.records.copy(deep=True)


def _nano_records():
    return pd.DataFrame(
        {
            "fif_childrace___1": [1, 1, 0],
            "fif_childrace___2": [0, 0, 0],
            "fif_childrace___3": [0, 0, 0],
            "fif_childrace___4": [0, 0, 0],
            "fif_childrace___5": [0, 0, 0],
            "fif_childrace___6": [0, 0, 0],
            "fif_childethnicity": [1, 2, pd.NA],
            "demo_ineligible": [1, 1, 0],
            "demo_unenrolled": [0, 0, 0],
            "demo_exclude": [0, 0, 0],
            "visit_date": ["2026-01-01", "2026-02-01", "bad-date"],
            "bsrc_doe": [pd.NA, pd.NA, pd.NA],
            "papf_parent_date": [pd.NA, pd.NA, pd.NA],
            "fif_doe": [pd.NA, pd.NA, pd.NA],
        },
        index=pd.Index(["n1", "n1", "n2"], name="demo_id"),
    )


def _nico_records():
    return pd.DataFrame(
        {
            "race___1": [1, 0],
            "race___2": [0, 0],
            "race___3": [0, 0],
            "race___4": [0, 1],
            "race___5": [0, 0],
            "race___6": [0, 0],
            "race___7": [0, 0],
            "fif_childethnicity": [2, pd.NA],
            "ethnicity": [1, pd.NA],
            "demo_ineligible": [0, 0],
            "demo_unenrolled": [0, 0],
            "demo_exclude": [0, 0],
            "dual_enrolled": [1, 0],
            "visit_date": [pd.NA, "2026-02-10"],
            "dob": ["2025-10-01", "2026-01-01"],
        },
        index=pd.Index(["c1", "c2"], name="id"),
    )


def _factory(api_url, token, *, timeout):
    assert api_url == "https://example.test/api/"
    assert timeout == (10, 90)
    if token == NANO_TOKEN:
        return FakeProject(STUDIES[0], _nano_records())
    if token == NICO_TOKEN:
        return FakeProject(STUDIES[1], _nico_records())
    raise AssertionError("unexpected token")


def test_load_env_file_does_not_override_existing_value(tmp_path, monkeypatch):
    env_file = tmp_path / ".env"
    env_file.write_text(
        "NANO_API_TOKEN=from-file\nNICO_API_TOKEN='from-file-two'\n",
        encoding="utf-8",
    )
    monkeypatch.setenv("NANO_API_TOKEN", "from-environment")
    monkeypatch.delenv("NICO_API_TOKEN", raising=False)
    load_env_file(env_file)
    assert os.environ["NANO_API_TOKEN"] == "from-environment"
    assert os.environ["NICO_API_TOKEN"] == "from-file-two"


def test_classification_preserves_overlapping_review_reasons():
    config = STUDIES[0]
    records = _nano_records().reset_index()
    audit, coverage = classify_records(
        config=config,
        records=records,
        metadata_fields=set(_metadata(config).reset_index()["field_name"]),
    )
    first = audit.loc[audit["demo_id"] == "n1"].iloc[0]
    second = audit.loc[audit["demo_id"] == "n2"].iloc[0]
    assert first["decision"] == "flagged-review"
    assert first["in_cumulative"] == 1
    assert first["flag_demo_ineligible"] == 1
    assert first["conflict_fif_childethnicity"] == 1
    assert "demo_ineligible=Yes" in first["reason"]
    assert "conflicting fif_childethnicity values" in first["reason"]
    assert second["missing_race"] == 1
    assert second["missing_ethnicity"] == 1
    assert "missing race" in second["reason"]
    visit = next(item for item in coverage if item["Field"] == "visit_date")
    assert visit["Participants with valid date"] == 1
    assert visit["Participants with invalid date"] == 1
    assert visit["Participants with multiple dates"] == 1


def test_strict_eligibility_excludes_flagged_record():
    config = STUDIES[0]
    audit, _ = classify_records(
        config=config,
        records=_nano_records().reset_index(),
        metadata_fields=set(_metadata(config).reset_index()["field_name"]),
        strict_eligibility=True,
    )
    first = audit.loc[audit["demo_id"] == "n1"].iloc[0]
    assert first["decision"] == "excluded"
    assert first["in_cumulative"] == 0


def test_record_ids_preserve_leading_zeroes_and_na_like_text():
    config = STUDIES[0]
    records = _nano_records().iloc[[0, 2]].reset_index(drop=True)
    records.insert(0, "demo_id", ["001", "NA"])
    audit, _ = classify_records(
        config=config,
        records=records,
        metadata_fields=set(_metadata(config).reset_index()["field_name"]),
    )
    assert audit["demo_id"].tolist() == ["001", "NA"]


def test_public_ratio_and_status_do_not_reveal_suppressed_actual():
    assert _public_ratio(7, 10) is None
    assert _public_status(7, 10) == "Suppressed"


def test_historical_cumulative_series_do_not_expose_small_exact_intervals():
    for values in STUDIES[0].historical_actuals.values():
        for previous, current in zip(values, values[1:]):
            if isinstance(previous, int) and isinstance(current, int):
                difference = current - previous
                assert difference == 0 or difference >= 10


def test_generate_reports_keeps_live_counts_out_of_historical_milestones(
    tmp_path, monkeypatch, capsys
):
    monkeypatch.setenv("NANO_API_TOKEN", NANO_TOKEN)
    monkeypatch.setenv("NICO_API_TOKEN", NICO_TOKEN)
    public = tmp_path / "public"
    secure = tmp_path / "secure"
    legacy = tmp_path / "legacy"
    result = generate_reports(
        report_date=date(2026, 7, 23),
        public_dir=public,
        secure_dir=secure,
        legacy_nano_dir=legacy,
        env_file=tmp_path / "missing.env",
        api_url="https://example.test/api/",
        project_factory=_factory,
    )

    assert result.studies["NANO"].live_counts == {
        "Total": 2,
        "Minority": 1,
        "Hispanic": 0,
    }
    assert result.studies["NICO"].live_counts == {
        "Total": 2,
        "Minority": 1,
        "Hispanic": 1,
    }
    _, nico_rows = report_rows(result.studies["NICO"], result.report_date)
    nico_actual = next(
        row for row in nico_rows if row["category"] == "Total" and row["kind"] == "actual"
    )
    assert all(value is None for value in nico_actual["values"])
    assert nico_actual["live"] == 2

    manifest = json.loads(result.paths["manifest"].read_text(encoding="utf-8"))
    assert manifest["projects"]["NICO"]["live_counts"]["Total"] == "<10"
    assert manifest["projects"]["NICO"]["live_counts"]["Hispanic"] == "<10"
    assert (
        manifest["projects"]["NICO"]["data_quality"]["Dual-enrolled in NANO"]
        == "<10"
    )
    assert "Strict-policy sensitivity total" not in manifest["projects"]["NANO"][
        "data_quality"
    ]
    assert "token" not in result.paths["manifest"].read_text(encoding="utf-8").lower()
    assert stat.S_IMODE(secure.stat().st_mode) == 0o700
    assert stat.S_IMODE(result.paths["nano_audit"].stat().st_mode) == 0o600
    assert stat.S_IMODE(result.paths["nico_audit"].stat().st_mode) == 0o600
    assert stat.S_IMODE(result.paths["secure_summary"].stat().st_mode) == 0o600
    secure_summary = json.loads(
        result.paths["secure_summary"].read_text(encoding="utf-8")
    )
    assert secure_summary["classification"] == "restricted-local"
    assert secure_summary["projects"]["NICO"]["live_counts"]["Hispanic"] == 1

    workbook = load_workbook(result.paths["workbook"], data_only=False)
    assert {"NANO", "NICO", "Run Summary", "Data Quality", "Date Candidates", "Provenance", "QA Checks"} <= set(workbook.sheetnames)
    nico = workbook["NICO"]
    assert nico.column_dimensions["A"].width >= 50
    assert nico.freeze_panes == "B5"
    nano = workbook["NANO"]
    ratio_row = next(
        row
        for row in range(1, nano.max_row + 1)
        if nano.cell(row, 1).value == "Actual / Current Target: Total Recruitment"
    )
    status_row = next(
        row
        for row in range(1, nano.max_row + 1)
        if nano.cell(row, 1).value == "Status: Total Recruitment"
    )
    assert nano.cell(ratio_row, 2).value is None
    assert nano.cell(ratio_row, 2).number_format == "0%"
    assert nano.cell(status_row, 2).value == "N/A — target unverified"
    nico_live_hispanic_row = next(
        row
        for row in range(1, nico.max_row + 1)
        if nico.cell(row, 1).value
        == "Published Historical Actual / Live-as-of Count: Hispanic Ethnicity Recruitment"
    )
    assert nico.cell(nico_live_hispanic_row, nico.max_column).value == "<10"

    assert (legacy / "nano_recruitment_milestones.html").is_file()
    assert (legacy / "nano_recruitment_milestones.xlsx").is_file()
    assert result.paths["nano_html"].read_text(encoding="utf-8").count("Live as of") >= 1
    console = capsys.readouterr().out
    assert "[NICO]" in console
    assert "Hispanic=<10" in console
    assert "Hispanic=1" not in console


def test_project_identity_mismatch_is_rejected(tmp_path, monkeypatch):
    monkeypatch.setenv("NANO_API_TOKEN", NANO_TOKEN)
    monkeypatch.setenv("NICO_API_TOKEN", NICO_TOKEN)

    def wrong_factory(api_url, token, *, timeout):
        if token == NANO_TOKEN:
            return FakeProject(STUDIES[0], _nano_records(), project_id=9999)
        return FakeProject(STUDIES[1], _nico_records())

    with pytest.raises(RuntimeError, match="unexpected REDCap project"):
        generate_reports(
            report_date=date(2026, 7, 23),
            public_dir=tmp_path / "public",
            secure_dir=tmp_path / "secure",
            legacy_nano_dir=tmp_path / "legacy",
            env_file=tmp_path / "missing.env",
            api_url="https://example.test/api/",
            project_factory=wrong_factory,
        )


def test_missing_configured_status_field_is_rejected(tmp_path, monkeypatch):
    monkeypatch.setenv("NANO_API_TOKEN", NANO_TOKEN)
    monkeypatch.setenv("NICO_API_TOKEN", NICO_TOKEN)

    class MissingStatusProject(FakeProject):
        def export_metadata(self, *, format_type):
            metadata = super().export_metadata(format_type=format_type)
            return metadata.drop(index=self.config.status_fields[0])

    def incomplete_factory(api_url, token, *, timeout):
        if token == NANO_TOKEN:
            return MissingStatusProject(STUDIES[0], _nano_records())
        return FakeProject(STUDIES[1], _nico_records())

    with pytest.raises(RuntimeError, match="missing required fields"):
        generate_reports(
            report_date=date(2026, 7, 23),
            public_dir=tmp_path / "public",
            secure_dir=tmp_path / "secure",
            legacy_nano_dir=tmp_path / "legacy",
            env_file=tmp_path / "missing.env",
            api_url="https://example.test/api/",
            project_factory=incomplete_factory,
        )


def test_formatted_export_failure_traceback_does_not_retain_token(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setenv("NANO_API_TOKEN", NANO_TOKEN)
    monkeypatch.setenv("NICO_API_TOKEN", NICO_TOKEN)

    def token_bearing_failure(api_url, token, *, timeout):
        raise RuntimeError(f"request rejected for token={token}")

    with pytest.raises(RuntimeError) as caught:
        generate_reports(
            report_date=date(2026, 7, 23),
            public_dir=tmp_path / "public",
            secure_dir=tmp_path / "secure",
            legacy_nano_dir=tmp_path / "legacy",
            env_file=tmp_path / "missing.env",
            api_url="https://example.test/api/",
            project_factory=token_bearing_failure,
        )

    formatted = "".join(
        traceback.format_exception(
            type(caught.value),
            caught.value,
            caught.value.__traceback__,
        )
    )
    assert NANO_TOKEN not in formatted
    assert "[redacted]" in formatted
